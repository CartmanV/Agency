#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build.py — сборка данных сайта направления «Агентства».

Итерация 0: читает единый бэклог v5.xlsx (вкладка «Бэклог», 219 итераций),
пересчитывает RICE и Final Score, валидирует и пишет site/data/backlog.json.

Принцип (см. «Сайт направления — архитектура и план», р. 3.1):
данные отдельно от представления. HTML не хранит чисел — читает data/*.json.

Запуск:  cd site && python build/build.py
"""

import json
import re
import sys
import unicodedata
from pathlib import Path
from urllib.parse import quote

try:
    import openpyxl
except ImportError:
    sys.exit("Нужен openpyxl:  pip3 install openpyxl")

try:
    import yaml
except ImportError:
    sys.exit("Нужен PyYAML:  pip3 install PyYAML")

# --- пути ---------------------------------------------------------------
SITE_DIR = Path(__file__).resolve().parent.parent          # site/
WORK_DIR = SITE_DIR.parent                                 # My work/  (источники)
DATA_DIR = SITE_DIR / "data"
SHEET = "Бэклог"

# --- машинный канон скоринга -------------------------------------------
# Единственный источник порогов, формул и кодов гипотез — общий с ботом файл.
# Прозаический канон: «Проектные инструкции v2.md», §5.3–5.4.
# Раньше эти правила дублировались здесь в коде; копия разъезжалась молча.
SCORING_PATH = WORK_DIR / "tg-bot-pm" / "config" / "scoring.yml"


def load_scoring():
    if not SCORING_PATH.exists():
        sys.exit(f"Не найден машинный канон скоринга: {SCORING_PATH}")
    with SCORING_PATH.open(encoding="utf-8") as fh:
        return yaml.safe_load(fh)


SCORING = load_scoring()
BOUNDS = SCORING["bounds"]
HYPOTHESIS_IDS = {h["id"] for group in SCORING["hypotheses"].values() for h in group}
MOSCOW_VALUES = set(SCORING["moscow_order"])
GATE_VALUES = set(SCORING["gate_values"])

# --- машинный канон таксономии -----------------------------------------
# §4 (подцели) и §5.1–5.2 (этапы, блоки, механизмы) живут в taxonomy.yml —
# том же файле, который читает бот. Второй копии этих словарей быть не должно:
# scoring.yml несёт только §5.3–5.4.
TAXONOMY_PATH = WORK_DIR / "tg-bot-pm" / "config" / "taxonomy.yml"


def load_taxonomy():
    if not TAXONOMY_PATH.exists():
        sys.exit(f"Не найден машинный канон таксономии: {TAXONOMY_PATH}")
    with TAXONOMY_PATH.open(encoding="utf-8") as fh:
        return yaml.safe_load(fh)


TAXONOMY = load_taxonomy()

# Написания «вне канона»: значение опознано, но записано не канонически.
# Это не ошибка данных, а материал для нормализации, поэтому копится
# отдельно от errors и печатается своим блоком.
STYLE_NOTES = []

# Осознанные gap'ы гипотез (§5.3 + решение Влада, список в scoring.yml).
# Считаем отдельно: gap должен быть виден как gap, а не пропасть в тишине.
HYPOTHESIS_GAPS = SCORING.get("hypothesis_gaps") or []
GAP_NOTES = []

# Пояснение в скобках: «топ-3 (52% IBC, 13 аг)», «1 «Освободить время» (основная)».
# Это комментарий к значению, а не другое значение — снимается ДО разбиения
# составных, иначе «1 + 2 (трудозатраты + тормоза)» рвётся по плюсу внутри скобок.
_PARENS = re.compile(r"[\(\[][^)\]]*[\)\]]")


def _strip_parens(s):
    s = unicodedata.normalize("NFC", str(s))
    prev = None
    while prev != s:                     # скобки бывают вложенными
        prev = s
        s = _PARENS.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def _norm_token(s):
    """Единый вид для сверки со словарём: NFC, без скобок, без кавычек-ёлочек."""
    s = _strip_parens(s)
    for q in "«»„“”\"'":
        s = s.replace(q, "")
    return re.sub(r"\s+", " ", s).strip(" .;:—-").strip()


def _split_top_level(text, splitters):
    """Разбить по разделителям, НЕ трогая то, что внутри скобок.

    «1 «Освободить время» (основная) + 2 (тормоза)» → два значения, а не четыре:
    плюс внутри пояснения разделителем не является.
    """
    parts, buf, depth = [], [], 0
    for ch in unicodedata.normalize("NFC", str(text)):
        if ch in "([":
            depth += 1
        elif ch in ")]":
            depth = max(0, depth - 1)
        if depth == 0 and ch in splitters:
            parts.append("".join(buf))
            buf = []
        else:
            buf.append(ch)
    parts.append("".join(buf))
    return [p.strip() for p in parts if p.strip()]


PRIMARY_MARKERS = [m.casefold() for m in (TAXONOMY.get("primary_markers") or [])]
SECONDARY_MARKERS = [m.casefold() for m in (TAXONOMY.get("secondary_markers") or [])]


def _role_of(part):
    """«Снимать нагрузку (осн.)» → 'primary'; «(вторичн.)» → 'secondary'; иначе None.

    Смотрим только в скобки — там канон и пишет пометку (см. карточки L1–L2).
    """
    tail = " ".join(_PARENS.findall(str(part))).casefold()
    if any(m in tail for m in PRIMARY_MARKERS):
        return "primary"
    if any(m in tail for m in SECONDARY_MARKERS):
        return "secondary"
    return None


def _index_from_entries(entries, canon_key="name"):
    """{нормализованное написание: (канон, это_алиас)} из списка taxonomy.yml.

    В бэклоге одно и то же поле пишут тремя способами сразу: «2», «Снимать
    нагрузку», «2 «Снимать нагрузку»». Все три должны опознаваться, иначе
    валидатор ругается на оформление, а не на смысл.

    canon_key — что считать эталонным написанием. Для подцели это `id`:
    §6 (шаблон карточки) требует номер, «Подцель направления: [1-7]».
    Для этапа и механизма — `name`.
    """
    idx = {}
    for e in entries:
        canon = str(e[canon_key])
        idx[_norm_token(canon).casefold()] = (canon, False)
        vid = e.get("id")
        name = str(e.get("name", ""))
        forms = [name] + list(e.get("aliases", []))
        if vid is not None:
            forms.append(str(vid))
            forms += [f"{vid} {f}" for f in [name] + list(e.get("aliases", [])) if f]
        for f in forms:
            if not f:
                continue
            idx.setdefault(_norm_token(f).casefold(), (canon, True))
    return idx


def _mechanism_index():
    idx = {}
    for name in TAXONOMY.get("mechanisms", []):
        idx[_norm_token(name).casefold()] = (str(name), False)
    for alias, canon in (TAXONOMY.get("mechanism_aliases") or {}).items():
        idx.setdefault(_norm_token(alias).casefold(), (canon, True))
    return idx


# Этапы 2A и 2B — один этап «Снимать нагрузку» с разными блоками, поэтому
# по имени они схлопываются; для валидации поля «Этап» это и нужно.
VOCAB_INDEX = {
    "stage": _index_from_entries(TAXONOMY.get("stages", [])),
    "block": _index_from_entries(TAXONOMY.get("blocks", [])),
    "subgoal": _index_from_entries(TAXONOMY.get("subgoals", []), canon_key="id"),
    "mechanism": _mechanism_index(),
}
SINGLE_VALUE = TAXONOMY.get("single_value", {})
SPLITTERS = TAXONOMY.get("splitters") or ["+", ",", "·"]

# Токены, законные в поле, но не являющиеся значением словаря:
# «функциональный паритет» — условие этапа 1 (§5.1), а не восьмая подцель.
EXTRA_TOKENS = {name: {_norm_token(t).casefold()
                       for t in (TAXONOMY.get(f"{name}_extra_tokens") or [])}
                for name in ("stage", "block", "mechanism", "subgoal")}
# «—» внутри составного значения — это «здесь пусто», а не неизвестное значение
# (встречается как «— / Сокращение трудозатрат», «— ; вторичный эффект …»).
_EMPTY_TOKENS = {"", "-", "–", "—", "н/д", "нд", "n/a"}


def _conc_index():
    """Метка «Концентрация»: словарь обеих осей + прежние написания (§5.4)."""
    spec = SCORING.get("concentration") or {}
    idx = {}
    for v in list(spec.get("volume_values") or []) + list(spec.get("money_values") or []):
        idx[_norm_token(v).casefold()] = (v, False)
    for old, canon in (spec.get("aliases") or {}).items():
        idx.setdefault(_norm_token(old).casefold(), (canon, True))
    return idx


CONC_INDEX = _conc_index()
# «деньги: ГСП+Космос» — шаблон денежной оси, конкретные агентства не словарные.
_CONC_MONEY_FREEFORM = re.compile(r"^деньги\s*:", re.IGNORECASE)


def check_concentration(rec, rnum, rec_id, errors):
    """Метка широты базы. Две оси через слэш: «объём / деньги» (§5.4)."""
    raw = rec.get("concentration")
    if raw is None or not str(raw).strip() or str(raw).strip() in ("—", "-", "–"):
        return
    if str(raw).rstrip().endswith("…"):
        return          # см. check_truncated
    text = _strip_parens(raw)
    if not text:
        return
    # «н/д» само содержит слэш — по нему оси не режем, иначе получаем «н» и «д».
    if _norm_token(text).casefold() in ("н/д", "нд", "n/a"):
        return
    for part in [p.strip() for p in text.split("/") if p.strip()]:
        if _CONC_MONEY_FREEFORM.match(part):
            continue                       # «деньги: <агентства>» — свободный список
        hit = CONC_INDEX.get(_norm_token(part).casefold())
        if hit is None:
            short = part if len(part) <= 45 else part[:42] + "…"
            errors.append(
                f"строка {rnum} ({rec_id}): Концентрация «{short}» — нет в словаре §5.4")
        elif hit[1]:
            STYLE_NOTES.append(
                f"строка {rnum} ({rec_id}): Концентрация «{part}» → канон «{hit[0]}»")


def check_vocab(rec, rnum, rec_id, errors, field, vocab_name, label):
    """Сверка поля со словарём §5.1 / §5.2 / §4.

    Три класса исхода:
      • значение не опознано               → ошибка (в errors);
      • составное там, где канон требует одного → ошибка (§5.1 про этап);
      • опознано, но написано не канонически   → STYLE_NOTES.
    Пустое значение здесь не трогаем: пустую подцель ловит отдельная проверка,
    а пустой механизм законен на этапе 1.
    """
    idx = VOCAB_INDEX.get(vocab_name) or {}
    if not idx:
        return
    raw = rec.get(field)
    if raw is None or not str(raw).strip() or str(raw).strip() in ("—", "-", "–"):
        return
    if str(raw).rstrip().endswith("…"):
        return          # обрезанную ячейку уже сообщил check_truncated, не дублируем

    # Разбиваем по верхнему уровню: плюс внутри скобок — часть пояснения,
    # а не разделитель значений.
    text = unicodedata.normalize("NFC", str(raw)).strip()
    parts = _split_top_level(text, SPLITTERS)
    if not parts:
        return

    mode = SINGLE_VALUE.get(vocab_name)

    # §5.1 (ред. 2026-08-06): у поля «Этап» ровно один ОСНОВНОЙ этап,
    # вторичный допускается с явной пометкой «(вторичн.)».
    if mode == "primary_marked" and len(parts) > 1:
        roles = [_role_of(p) for p in parts]
        primaries = roles.count("primary")
        if primaries == 0:
            errors.append(
                f"строка {rnum} ({rec_id}): {label} «{_strip_parens(text)}» — "
                f"несколько этапов, ни один не помечен основным "
                f"(§5.1: «Снимать нагрузку (осн.) + Не мешать (вторичн.)»)")
            return
        if primaries > 1:
            errors.append(
                f"строка {rnum} ({rec_id}): {label} «{_strip_parens(text)}» — "
                f"основных этапов {primaries}, а канон требует ровно одного (§5.1)")
            return
    elif mode is True and len(parts) > 1:
        errors.append(
            f"строка {rnum} ({rec_id}): {label} «{_strip_parens(text)}» — составное "
            f"значение, а канон требует ровно одного")
        return

    unknown, non_canon = [], []
    extra = EXTRA_TOKENS.get(vocab_name, set())
    for p in parts:
        key = _norm_token(p).casefold()
        if key in _EMPTY_TOKENS or key in extra:
            continue                       # пусто или законный не-словарный токен
        hit = idx.get(key)
        if hit is None:
            unknown.append(_strip_parens(p) or p)
        elif hit[1]:
            non_canon.append((_strip_parens(p) or p, hit[0]))

    if unknown:
        short = [u if len(u) <= 45 else u[:42] + "…" for u in unknown]
        errors.append(
            f"строка {rnum} ({rec_id}): {label} {short} — нет в словаре scoring.yml")
    for got, canon in non_canon:
        STYLE_NOTES.append(
            f"строка {rnum} ({rec_id}): {label} «{got}» → канон «{canon}»")


# Ячейка, обрезанная при записи. Найдено 2026-08-06: 62 ячейки в двух бэклогах
# кончаются на «…» ровно на 121 / 180 / 181 символе — след скрипта, писавшего
# с ограничением длины. Текст потерян в источнике, не при чтении, поэтому это
# ошибка данных, а не оформления: «7 «Миним…» уже не восстановить из файла.
_TRUNCATED_FIELDS = (("hypothesis", "Гипотеза"), ("subgoal", "Подцель"),
                     ("mechanism", "Механизм"), ("stage", "Этап"),
                     ("block", "Блок"), ("concentration", "Концентрация"))


def check_truncated(rec, rnum, rec_id, errors):
    for field, label in _TRUNCATED_FIELDS:
        val = rec.get(field)
        if isinstance(val, str) and val.rstrip().endswith("…"):
            errors.append(
                f"строка {rnum} ({rec_id}): {label} обрезана при записи "
                f"({len(val)} симв., кончается на «…») — текст потерян в источнике")


# Связь «этап ↔ подцель» из §4/§5.1 живёт в taxonomy.yml с самого начала
# (поля stages[].subgoals и subgoals[].stage), но никто её не проверял.
# Находка 2026-08-07: тему «Предложения 2.0» перепривязали с этапа 2B на этап 5,
# а подцель осталась прежняя — 31 строка несёт «этап 5 + подцель 2».
# Перепривязка сделана наполовину, и это было не видно.
# Берём расширенные наборы из stage_subgoals_allowed: stages[].subgoals — это
# ОСНОВНАЯ подцель этапа, а §4.1 разрешает подцелям 1, 2 и 7 жить на обоих
# этапах первичного контура. Без этого валидатор ругался на сам канон.
STAGE_SUBGOALS = {k: {str(x) for x in v}
                  for k, v in (TAXONOMY.get("stage_subgoals_allowed") or {}).items()}
for _st in TAXONOMY.get("stages", []):
    STAGE_SUBGOALS.setdefault(str(_st["name"]), set()).update(
        str(x) for x in (_st.get("subgoals") or []))
SUBGOAL_IDS = {str(g["id"]) for g in TAXONOMY.get("subgoals", [])}


def check_stage_subgoal(rec, rnum, rec_id, errors):
    """Основной этап и подцели должны сходиться по таблице §4."""
    stage_raw, sub_raw = rec.get("stage"), rec.get("subgoal")
    if not stage_raw or not sub_raw:
        return
    if str(stage_raw).rstrip().endswith("…") or str(sub_raw).rstrip().endswith("…"):
        return

    # Основной этап: помеченный «(осн.)», иначе единственный.
    parts = _split_top_level(str(stage_raw), SPLITTERS)
    primary = next((p for p in parts if _role_of(p) == "primary"), parts[0] if parts else None)
    if not primary:
        return
    hit = VOCAB_INDEX["stage"].get(_norm_token(primary).casefold())
    if not hit:
        return
    allowed = STAGE_SUBGOALS.get(hit[0])
    if not allowed:
        return

    subs = set()
    for p in _split_top_level(str(sub_raw), SPLITTERS):
        h = VOCAB_INDEX["subgoal"].get(_norm_token(p).casefold())
        if h:
            subs.add(h[0])
    if not subs or (subs & allowed):
        return          # хотя бы одна подцель соответствует этапу — норма
    errors.append(
        f"строка {rnum} ({rec_id}): этап «{hit[0]}» ждёт подцель "
        f"{sorted(allowed)}, а стоит {sorted(subs)} — связь §4 не сходится")


def check_scoring(rec, rnum, rec_id, errors, with_ml=False):
    """Валидация строки по scoring.yml. Не правит данные — только сообщает."""
    for key, field in (("reach", "reach"), ("impact", "impact"),
                       ("confidence", "confidence"), ("pv_mult", "pvMult")):
        val = rec.get(field)
        if val is None or not isinstance(val, (int, float)):
            continue
        lo, hi = BOUNDS[key]
        if not (lo <= val <= hi):
            errors.append(f"строка {rnum} ({rec_id}): {field}={val} вне канона [{lo}; {hi}]")
    if with_ml:
        for key, field in (("ml", "mlLeverage"), ("depth", "depth")):
            val = rec.get(field)
            if isinstance(val, (int, float)):
                lo, hi = BOUNDS[key]
                if not (lo <= val <= hi):
                    errors.append(f"строка {rnum} ({rec_id}): {field}={val} вне диапазона [{lo}; {hi}]")
    eff = rec.get("effort")
    if isinstance(eff, (int, float)) and eff <= 0:
        errors.append(f"строка {rnum} ({rec_id}): Effort={eff} — должен быть > 0")
    mos = rec.get("moscow")
    if mos and mos not in MOSCOW_VALUES:
        errors.append(f"строка {rnum} ({rec_id}): MoSCoW {mos!r} не из канона {sorted(MOSCOW_VALUES)}")
    gate = rec.get("gate")
    if gate and gate not in GATE_VALUES:
        errors.append(f"строка {rnum} ({rec_id}): Gate {gate!r} не из канона {sorted(GATE_VALUES)}")
    hyp = rec.get("hypothesis")
    if hyp:
        text = str(hyp).strip()
        codes = re.findall(r"H\d+\.\d+", text)
        unknown = [c for c in codes if c not in HYPOTHESIS_IDS]
        if unknown:
            errors.append(f"строка {rnum} ({rec_id}): гипотеза {unknown} нет в scoring.yml")
        # Плейсхолдер вместо кода. До 2026-08-06 сюда проваливались 49 строк:
        # «H5.x — платный сервис поверх (уточнить код)» не даёт совпадений по
        # H\d+\.\d+, значит и ошибки не давало. §5.3 требует ЯВНОЙ ссылки на код.
        elif not codes:
            gap = next((g for g in HYPOTHESIS_GAPS
                        if text.casefold().startswith(str(g["id"]).casefold())), None)
            if gap:
                # Осознанный gap (решение Влада, зафиксировано в scoring.yml).
                # Не ошибка, но и не тишина: считаем отдельно, чтобы было видно.
                GAP_NOTES.append(f"строка {rnum} ({rec_id}): {gap['id']} — {gap['scope']}")
            else:
                short = text if len(text) <= 60 else text[:57] + "…"
                errors.append(
                    f"строка {rnum} ({rec_id}): гипотеза «{short}» — без кода H1.x–H5.x "
                    f"(§5.3 требует явной ссылки)")

    # Словари §5.1 / §5.2 / §4 — до 2026-08-06 эти поля не проверялись вовсе.
    check_vocab(rec, rnum, rec_id, errors, "stage", "stage", "Этап")
    check_vocab(rec, rnum, rec_id, errors, "block", "block", "Блок")
    check_vocab(rec, rnum, rec_id, errors, "mechanism", "mechanism", "Механизм")
    check_vocab(rec, rnum, rec_id, errors, "subgoal", "subgoal", "Подцель")
    check_concentration(rec, rnum, rec_id, errors)
    check_truncated(rec, rnum, rec_id, errors)
    check_stage_subgoal(rec, rnum, rec_id, errors)


def build_meta():
    """Штамп сборки — чтобы на странице было видно, насколько свежи данные."""
    from datetime import datetime, timezone
    return {
        "builtAt": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "scoringSynced": SCORING.get("synced"),
        "scoringSource": SCORING.get("source"),
    }

# Заголовок колонки xlsx -> ключ в JSON
COLUMNS = {
    "#": "num",
    "Статус": "status",
    "MoSCoW": "moscow",
    "Тема": "theme",
    "Уровень 2": "level2",
    "Итерация": "iteration",
    "ID гитлаб": "gitlabId",
    "Название": "title",
    "Проблема (Job Story)": "jobStory",
    "Проблема: источник": "problemSource",
    "Этап": "stage",
    "Блок": "block",
    "Механизм": "mechanism",
    "Подцель": "subgoal",
    "Гипотеза": "hypothesis",
    "Reach": "reach",
    "Impact": "impact",
    "Confidence": "confidence",
    "Effort": "effort",
    "RICE Score": "rice",
    "PV Mult": "pvMult",
    "Final Score": "finalScore",
    "Gate": "gate",
    "Концентрация": "concentration",
    "Цитата / данные": "quote",
    "Агентства (исследования)": "agencies",
    "Обоснование": "rationale",
    "PV Notes": "pvNotes",
    "Активные метрики": "activeMetrics",
    "Целевые метрики": "targetMetrics",
}

NUMERIC = {"num", "reach", "impact", "confidence", "effort", "pvMult"}


def slug(value):
    """Транслитерация в стабильный ascii-slug для id (тема/уровень)."""
    if value is None:
        return ""
    table = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    }
    s = "".join(table.get(ch, ch) for ch in str(value).lower())
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


# Единый термин: «операция» (как единица работы агентства) → «транзакция».
# \b защищает «операционные/операционка» и «кооперация» (другая основа). Применяется
# к собранным data/*.json после сборки, чтобы термин был устойчив к пересборке.
_TERM_RX = re.compile(r"\b(операци|Операци)(я|и|й|ю|ей|ях|ями|ям)\b")


def _term_repl(m):
    return ("Транзакци" if m.group(1)[0] == "О" else "транзакци") + m.group(2)


def normalize_terms():
    total = 0
    for p in sorted(DATA_DIR.glob("*.json")):
        text = p.read_text(encoding="utf-8")
        new, n = _TERM_RX.subn(_term_repl, text)
        if n:
            p.write_text(new, encoding="utf-8")
            total += n
    if total:
        print(f"Термин «операция→транзакция»: {total} замен в data/*.json")


def find_source():
    """Самый свежий файл единого бэклога (по версии vN) в папке My work.
    Устойчиво к ревизиям (v5 → v6 → …) и NFD/NFC-нормализации имён на macOS."""
    cands = []
    for p in WORK_DIR.glob("*.xlsx"):
        name = unicodedata.normalize("NFC", p.name)
        low = name.lower()
        if "бэклог направлени" in low and "соответствие" not in low and "черновик" not in low:
            m = re.search(r"\bv(\d+)\b", name)
            cands.append((int(m.group(1)) if m else 0, name, p))
    if not cands:
        sys.exit(f"Не найден файл единого бэклога (*.xlsx с «бэклог направления») в {WORK_DIR}")
    cands.sort(key=lambda c: (c[0], c[1]))
    return cands[-1][2]  # наивысшая версия


def build_backlog():
    src = find_source()
    print(f"Источник: {src.name}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Нет вкладки {SHEET!r}. Есть: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = [clean(h) for h in rows[0]]

    missing = [c for c in COLUMNS if c not in header]
    if missing:
        sys.exit(f"В источнике нет колонок: {missing}")
    idx = {c: header.index(c) for c in COLUMNS}

    items, errors, seen_ids = [], [], set()
    for rnum, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        rec = {}
        for col, key in COLUMNS.items():
            val = clean(row[idx[col]])
            if key in NUMERIC and val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (TypeError, ValueError):
                    pass
            rec[key] = val

        # --- пересчёт RICE и Final (формулы в xlsx не закешированы) ---
        r, i, c, e = rec.get("reach"), rec.get("impact"), rec.get("confidence"), rec.get("effort")
        pv = rec.get("pvMult")
        if None not in (r, i, c, e) and e:
            rec["rice"] = round((r * i * c) / e)
        else:
            rec["rice"] = None
        if rec["rice"] is not None and pv is not None:
            rec["finalScore"] = round(rec["rice"] * pv)
        else:
            rec["finalScore"] = None

        # --- стабильный id: <тема>-<уровень2>-<#> (повтор тема==L2 схлопываем) ---
        parts, prev = [], None
        for p in (slug(rec.get("theme")), slug(rec.get("level2")), str(rec.get("num"))):
            if p and p != prev:
                parts.append(p)
            prev = p
        rec_id = "-".join(parts)
        rec["id"] = rec_id
        rec["cardPath"] = None  # гибрид карточек L3 — позже (итерация 3)

        # --- валидация ---
        if rec_id in seen_ids:
            errors.append(f"строка {rnum}: дубль id {rec_id!r}")
        seen_ids.add(rec_id)
        if not rec.get("subgoal"):
            errors.append(f"строка {rnum} ({rec_id}): пустая подцель")
        if not rec.get("hypothesis"):
            errors.append(f"строка {rnum} ({rec_id}): пустая гипотеза")
        check_scoring(rec, rnum, rec_id, errors)

        items.append(rec)

    return items, errors


# --- Инициативы (сокращённый бэклог) ----------------------------------
# Отдельный источник: «Сокращённый бэклог — инициативы vN.xlsx», вкладка «Инициативы».
# Та же логика ценности/скоринга, но строки — укрупнённые инициативы (свёртки итераций):
# нет колонок «Цитата / данные» и «Агентства (исследования)», добавлены
# «Reach v1.5 — обоснование», «Задач в инициативе», «Свёрнутые итерации».
INIT_SHEET = "Инициативы"
INIT_COLUMNS = {
    "#": "num",
    "Статус": "status",
    "MoSCoW": "moscow",
    "Работа (JTBD, tree.html)": "jtbdWork",
    "ML (рычаг метрик)": "mlLeverage",
    "Depth (глубина)": "depth",
    "Метрики (сводно)": "metricsSummary",
    "Тема": "theme",
    "Уровень 2": "level2",
    "Итерация": "iteration",
    "ID гитлаб": "gitlabId",
    "Название": "title",
    "Проблема (Job Story)": "jobStory",
    "Проблема: источник": "problemSource",
    "Этап": "stage",
    "Блок": "block",
    "Механизм": "mechanism",
    "Подцель": "subgoal",
    "Гипотеза": "hypothesis",
    "Reach": "reach",
    "Impact": "impact",
    "Confidence": "confidence",
    "Effort": "effort",
    "RICE Score": "rice",
    "PV Mult": "pvMult",
    "Final Score": "finalScore",
    "Gate": "gate",
    "Концентрация": "concentration",
    "Обоснование": "rationale",
    "PV Notes": "pvNotes",
    "Активные метрики": "activeMetrics",
    "Целевые метрики": "targetMetrics",
    "Reach v1.5 — обоснование": "reachNote",
    "Задач в инициативе": "taskCount",
    "Свёрнутые итерации": "collapsedIterations",
}
INIT_NUMERIC = {"num", "reach", "impact", "confidence", "effort", "pvMult", "taskCount", "mlLeverage", "depth"}


def find_initiatives_source():
    """Самый свежий «Сокращённый бэклог — инициативы vN.xlsx» в папке My work
    (устойчиво к ревизиям v1 → v1.1 → … и NFD/NFC-нормализации имён на macOS)."""
    cands = []
    for p in WORK_DIR.glob("*.xlsx"):
        low = unicodedata.normalize("NFC", p.name).lower()
        if "сокращённый бэклог" in low and "инициатив" in low:
            m = re.search(r"\bv(\d+(?:\.\d+)*)", unicodedata.normalize("NFC", p.name))
            ver = tuple(int(x) for x in m.group(1).split(".")) if m else (0,)
            cands.append((ver, p.name, p))
    if not cands:
        return None
    cands.sort(key=lambda c: (c[0], c[1]))
    return cands[-1][2]


def build_initiatives():
    """Вкладка «Инициативы» → initiatives.json. Та же структура, что backlog.json
    (id, пересчёт RICE/Final), плюс поля taskCount/collapsedIterations/reachNote."""
    src = find_initiatives_source()
    if src is None:
        return None, ["источник инициатив (*сокращённый бэклог*инициатив*.xlsx) не найден"]
    print(f"Источник инициатив: {src.name}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    if INIT_SHEET not in wb.sheetnames:
        return None, [f"нет вкладки {INIT_SHEET!r} в {src.name}"]
    ws = wb[INIT_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header = [clean(h) for h in rows[0]]
    idx = {c: header.index(c) for c in INIT_COLUMNS if c in header}
    missing = [c for c in INIT_COLUMNS if c not in header]

    items, errors, seen_ids = [], list(missing and [f"нет колонок: {missing}"]), set()
    for rnum, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        rec = {}
        for col, key in INIT_COLUMNS.items():
            if col not in idx:
                rec[key] = None
                continue
            val = clean(row[idx[col]])
            if key in INIT_NUMERIC and val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (TypeError, ValueError):
                    pass
            rec[key] = val

        r, i, c, e = rec.get("reach"), rec.get("impact"), rec.get("confidence"), rec.get("effort")
        pv = rec.get("pvMult")
        rec["rice"] = round((r * i * c) / e) if None not in (r, i, c, e) and e else None
        rec["finalScore"] = round(rec["rice"] * pv) if rec["rice"] is not None and pv is not None else None

        # Модель Metric Leverage (v1.4): Value = I × C × ML × Depth × PV.
        # В xlsx это формула без кеша — считаем здесь; Norm% нормируем ниже по максимуму.
        ml, dp = rec.get("mlLeverage"), rec.get("depth")
        if None not in (i, c, ml, dp, pv):
            rec["value"] = round(i * c * ml * dp * pv, 1)
        else:
            rec["value"] = None

        parts, prev = [], None
        for p in (slug(rec.get("theme")), slug(rec.get("level2")), str(rec.get("num"))):
            if p and p != prev:
                parts.append(p)
            prev = p
        rec_id = "-".join(parts)
        rec["id"] = rec_id
        rec["cardPath"] = None

        if rec_id in seen_ids:
            errors.append(f"строка {rnum}: дубль id {rec_id!r}")
        seen_ids.add(rec_id)
        check_scoring(rec, rnum, rec_id, errors, with_ml=True)
        items.append(rec)

    # Norm % — доля от максимального Value (100% у сильнейшей инициативы).
    vmax = max((x["value"] for x in items if x.get("value") is not None), default=None)
    for x in items:
        x["normPct"] = round(x["value"] / vmax * 100) if x.get("value") is not None and vmax else None

    # Сквозная нумерация по приоритету: № = ранг в дефолтном порядке таблицы
    # (MoSCoW → Value убыв., как в app.js), пустые — вниз.
    # Номер из xlsx сохраняем как srcNum (он остаётся в id — ссылки стабильны).
    moscow_order = {"Must": 0, "Should": 1, "Could": 2, "Won't": 3}
    items.sort(key=lambda x: (moscow_order.get(x.get("moscow"), 9),
                              x.get("value") is None,
                              -(x.get("value") or 0),
                              -(x.get("finalScore") or 0),
                              str(x.get("title") or "")))
    for rank, x in enumerate(items, start=1):
        x["srcNum"] = x.get("num")
        x["num"] = rank

    return {"source": src.name, "count": len(items), "meta": build_meta(), "items": items}, errors


def build_tree(items):
    """Иерархия Тема (L1) → Уровень 2 → Итерация (ссылка на backlog id).
    Лестница и механизмы — оси-бейджи, не ветви (см. план р. 4.2)."""
    from collections import OrderedDict
    themes = OrderedDict()
    for it in items:
        th = it.get("theme") or "—"
        l2 = it.get("level2") or "—"
        themes.setdefault(th, OrderedDict()).setdefault(l2, []).append({
            "id": it["id"], "num": it["num"], "title": it.get("title"),
            "iteration": it.get("iteration"), "moscow": it.get("moscow"),
            "stage": it.get("stage"), "mechanism": it.get("mechanism"),
            "subgoal": it.get("subgoal"), "finalScore": it.get("finalScore"),
            "gate": it.get("gate"),
        })
    tree = []
    for th, l2map in themes.items():
        l2list, tcount, tmust = [], 0, 0
        for l2, arr in l2map.items():
            tcount += len(arr)
            tmust += sum(1 for x in arr if x["moscow"] == "Must")
            l2list.append({"name": l2, "count": len(arr), "items": arr})
        tree.append({"theme": th, "count": tcount, "mustCount": tmust, "level2": l2list})
    tree.sort(key=lambda t: -t["count"])
    return tree


def build_legend(src):
    """Вкладка «Легенда» (термин→определение, секции ВЕРХним регистром) → legend.json."""
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    if "Легенда" not in wb.sheetnames:
        return None
    ws = wb["Легенда"]
    rows = list(ws.iter_rows(values_only=True))

    # чистые id для известных секций (по ключевому слову заголовка)
    def section_id(title):
        t = title.lower()
        for key, sid in (("о файле", "about"), ("как читать", "columns"),
                         ("механизм", "mechanisms"), ("подцель", "subgoals"),
                         ("этап", "stages"), ("скоринг", "scoring"),
                         ("метрик", "metrics"), ("пристыков", "research")):
            if key in t:
                return sid
        return slug(title) or "section"

    sections, cur = [], None
    for r in rows[1:]:  # row0 — общий заголовок файла
        a = clean(r[0])
        b = clean(r[1]) if len(r) > 1 else None
        if a and not b:                       # заголовок секции
            cur = {"id": section_id(a), "title": a, "items": []}
            sections.append(cur)
        elif a and b and cur is not None:     # термин → определение
            cur["items"].append({"id": f"{cur['id']}--{slug(a)}", "term": a, "def": b})
    return {"source": src.name, "sections": sections,
            "count": sum(len(s["items"]) for s in sections)}


def find_agencies_source():
    """Файл «Свод по агентствам …xlsx» (устойчиво к NFD/NFC и дате)."""
    cands = [p for p in WORK_DIR.glob("*.xlsx")
             if "свод по агентствам" in unicodedata.normalize("NFC", p.name).lower()]
    return sorted(cands, key=lambda p: unicodedata.normalize("NFC", p.name))[-1] if cands else None


# Соответствие «потерянное агентство» → тип оттока (кураторская классификация, лист
# «3. Отток и миграции» даёт только сегмент Lost/Rising/…, не различает причину).
# Источник классификации — прежние анализы 2026-04 («Удержание и сегментация»).
CHURN_TYPE_MAP = {
    "TUI Corporate FS Travel, ООО \"ТТ-Т": "уход к конкуренту",
    "Да-Тревел": "уход к конкуренту",
    "ATPI": "уход к конкуренту",
    "Демлинк Атлас": "cross-agency миграция",
    "Броневик TMC": "cross-agency миграция",
    "Демо клиент": "прочее / демо",
    "OBT Raketa Travel": "прочее / демо",
    "Aeroglobus KZ": "прочее / демо",
}

# Клиентские помесячные листы «Общий отчёт …xlsx» (сырой транзакционный дамп: ID клиента ×
# месяц) — по одному листу на активное агентство. Единый свод больше не содержит истории
# клиентов (нет листа «Клиенты (NSM)»), поэтому NSM-историю разворачиваем из этого источника.
CLIENT_SHEET_MAP = {
    3156: "IBC. Помесячно",
    399: "ATH. Помесячно",
    32199: "Global Air. Помесячно",
    33121: "ГСП-Сервис. Помесячно",
    33504: "Интерсити Сервис (KMP Group)",
    32356: "Corporate travel. Помесячно",
    33261: "Aeroglobus. Помесячно",
    33021: "KazTour Corporate. Помесячно",
    32873: "Альбатрос. Помесячно",
    33310: "Космос Тревел. Помесячно",
    34500: "Симпл Флайт",
    33590: "Аэротон",
    34747: "Рейна-Тур",
    34515: "Атланта БТК",
    34723: "ИнТоп Консалтинг",
    34917: "FCMT",
    34774: "Иналекс",
    34899: "CTS-travel",
    34852: "Альянс Авиа",
}


def _anum(v):
    """Общий числовой парсер для листов свода агентств (None/"" → None)."""
    if v is None or v == "":
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def _resolve_agency_id(nm, agencies):
    """Короткое/неполное имя агентства (из вспомогательных листов) → id по подстроке в полном имени."""
    low = str(nm).lower()
    for a in agencies:
        full = str(a["name"]).lower()
        if low in full or full in low:
            return a["id"]
    return None


def build_agencies_clients(agencies):
    """История клиентов (NSM) по месяцам — из клиентских листов «Общий отчёт …xlsx» (ID клиента ×
    месяц на агентство). Единый свод («Свод по агентствам …xlsx») эту историю больше не содержит —
    разворачиваем активных клиентов из сырого дампа: клиент «активен» в месяце, если есть операции.
    «Сейчас»/«3 мес. назад» — скользящее окно 3 месяца (см. пояснение NSM на странице), не месяц день-в-день."""
    src = find_work_xlsx("общий отчет")
    if not src:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    per_agency, all_months = {}, set()
    for aid, sheet_name in CLIENT_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        it = wb[sheet_name].iter_rows(values_only=True)
        hdr = next(it)
        month_idx = [(i, str(h).strip()) for i, h in enumerate(hdr)
                     if h is not None and re.match(r"^\d{6}$", str(h).strip())]
        months = [m for _, m in month_idx]
        all_months.update(months)
        active = {m: set() for m in months}
        for r in it:
            cid = r[0]
            if cid is None:
                continue
            cid = str(cid)
            for i, m in month_idx:
                if i < len(r) and _anum(r[i]):
                    active[m].add(cid)
        per_agency[aid] = active

    months_sorted = sorted(all_months)   # "YYYYMM" — лексикографический порядок = хронологический
    if not months_sorted:
        return None
    latest_i = len(months_sorted) - 1

    def window(end_i, span):
        if end_i < 0:
            return []
        start_i = max(0, end_i - span + 1)
        return months_sorted[start_i:end_i + 1]

    def active_union(aid, months):
        active = per_agency.get(aid, {})
        s = set()
        for m in months:
            s |= active.get(m, set())
        return s

    now_w, ago_w, l6_w = window(latest_i, 3), window(latest_i - 3, 3), window(latest_i, 6)

    by_id, totals = {}, {"l6m": 0, "activeMo": 0, "ago": 0, "now": 0, "new": 0, "lost": 0}
    for a in agencies:
        aid = a["id"]
        now_s, ago_s, l6_s = active_union(aid, now_w), active_union(aid, ago_w), active_union(aid, l6_w)
        active_mo = len(per_agency.get(aid, {}).get(months_sorted[latest_i], set()))
        new_s, lost_s = now_s - ago_s, ago_s - now_s
        rec = {
            "clActiveMo": active_mo, "clNow": len(now_s), "clAgo": len(ago_s),
            "clNew": len(new_s), "clLost": len(lost_s), "clNet": len(now_s) - len(ago_s),
            "clL6m": len(l6_s),
        }
        by_id[aid] = rec
        for k in ("l6m", "activeMo", "ago", "now"):
            totals[k] += rec["cl" + k[0].upper() + k[1:]] if k != "l6m" else rec["clL6m"]
        totals["new"] += rec["clNew"]
        totals["lost"] += rec["clLost"]
    totals["net"] = totals["now"] - totals["ago"]

    # Помесячный ряд (до 13 точек, аналог ops L13M) — активных клиентов в месяц, всего и по id.
    win13 = months_sorted[-13:]
    by_id_series, total_series = {str(a["id"]): [] for a in agencies}, []
    for m in win13:
        tot = 0
        for a in agencies:
            n = len(per_agency.get(a["id"], {}).get(m, set()))
            by_id_series[str(a["id"])].append(n)
            tot += n
        total_series.append(tot)
    months_disp = [f"{m[4:6]}.{m[:4]}" for m in win13]

    return {
        "source": src.name,
        "nsm": totals,
        "byId": by_id,
        "monthlyClients": {"months": months_disp, "total": total_series, "byId": by_id_series},
    }


def build_agencies_abcdx(agencies):
    """ABCDX — сегментация клиентов по объёму (Парето: A ≤50% накопл. / B 50–80% / C 80–95% /
    D 95–100% / X ≤2 опер/мес). Источник — вкладки «1. Клиенты ABCDX» (сводно по буквам, БЕЗ
    имён клиентов — только агрегаты) и «Клиенты по агентствам (ABCDX)» (разбивка по агентствам)
    единого свода. Полный клиентский список (лист «1.…», с именами компаний) на сайт не идёт —
    приватность (см. CLAUDE.md)."""
    src = find_agencies_source()
    if not src:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    segments = []
    if "1. Клиенты ABCDX" in wb.sheetnames:
        started = False
        for r in wb["1. Клиенты ABCDX"].iter_rows(values_only=True):
            c0 = clean(r[0])
            if c0 == "Сегмент" and clean(r[1]) == "Клиентов":
                started = True; continue
            if not started:
                continue
            if c0 not in ("A", "B", "C", "D", "X"):
                break
            segments.append({"seg": c0, "clients": _anum(r[1]), "ops": _anum(r[2]), "opsPct": _anum(r[3])})

    by_agency, total = [], None
    sheet2 = "Клиенты по агентствам (ABCDX)"
    if sheet2 in wb.sheetnames:
        for r in list(wb[sheet2].iter_rows(values_only=True))[1:]:
            nm = clean(r[0])
            if not nm:
                continue
            rec = {"name": nm, "clients": _anum(r[1]), "ops": _anum(r[2]),
                   "a": _anum(r[3]), "b": _anum(r[4]), "c": _anum(r[5]),
                   "d": _anum(r[6]), "x": _anum(r[7])}
            if str(nm).startswith("ИТОГО"):
                total = rec; continue
            rec["id"] = _resolve_agency_id(nm, agencies)
            by_agency.append(rec)

    return {"segments": segments, "byAgency": by_agency, "total": total}


def build_agencies_price(agencies):
    """Ось ДЕНЕГ к ABCDX — лист «Цена транзакции» единого свода: тариф, оценка выручки и
    эффективная цена операции по агентствам + пересчёт концентрации на деньги.

    Три оговорки, которые обязаны доехать до страницы (иначе цифры читаются неверно):
      1. Выручка — РАСЧЁТ (тариф × операции июня), а не выгрузка из биллинга.
      2. IBC — внутри холдинга: тариф 170 ₽ считается по всем полученным от клиента деньгам,
         а не только по транзакционному сбору. С внешними 32–100 ₽ несопоставим → отдельная
         доля «без IBC» и флаг excluded.
      3. Аэроглобус — две цены по статусам, точной раскладки нет: выручка диапазоном.
    Источник цен — «Транзакции агентств. Определения и стоимость.xlsx» (2026); цены IBC 170 /
    KazTour 100 в том файле отсутствуют и внесены со слов PM (допущение)."""
    src = find_agencies_source()
    if not src or "Цена транзакции" not in openpyxl.load_workbook(
            src, read_only=True).sheetnames:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["Цена транзакции"]

    rows_out, totals, conc, mode = [], {}, [], None
    for r in ws.iter_rows(values_only=True):
        c0 = clean(r[0])
        if c0 == "Агентство" and clean(r[1]) == "Сегмент (Свод)":
            mode = "rows"; continue
        if c0 == "Срез" and clean(r[1]) == "По объёму":
            mode = "conc"; continue
        if c0 and str(c0).startswith(("C.", "D.", "E.", "F.", "G.")):
            mode = None
        if not c0 or not mode:
            continue
        if mode == "rows":
            if str(c0).startswith("ИТОГО"):
                totals["exIbc" if "без IBC" in str(c0) else "all"] = {
                    "ops": _anum(r[2]), "revenue": _anum(r[4]), "effPrice": _anum(r[5])}
                continue
            rows_out.append({
                "name": c0, "segment": clean(r[1]), "ops": _anum(r[2]),
                "tariff": clean(r[3]),
                "revenue": _anum(r[4]), "revenueText": None if _anum(r[4]) is not None else clean(r[4]),
                "effPrice": _anum(r[5]),
                "revSharePct": _anum(r[6]), "revShareExIbcPct": _anum(r[7]),
                "opsSharePct": _anum(r[8]), "supportRatio": _anum(r[9]),
                "offlinePct": _anum(r[10]), "flag": clean(r[11]),
                "excluded": c0 == "IBC",
                "id": _resolve_agency_id(c0, agencies),
            })
        elif mode == "conc":
            # после таблицы срезов идут прозаические строки-выводы (только колонка A) — не таблица
            if clean(r[1]) is None:
                mode = None; continue
            conc.append({"cut": c0, "byOps": clean(r[1]), "byMoney": clean(r[2]),
                         "byMoneyExIbc": clean(r[3]), "note": clean(r[4])})

    return {
        "rows": rows_out, "totals": totals, "concentration": conc,
        "priceSource": "Транзакции агентств. Определения и стоимость.xlsx (2026)",
        "assumption": ("Выручка — расчёт «тариф × операции июня», не выгрузка из биллинга. "
                       "Цены IBC 170 ₽ и KazTour 100 ₽ в файле цен отсутствуют — внесены со слов PM. "
                       "Экономика клиента IBC ≈ ×8 к клиенту внешнего агентства — оценка PM, "
                       "в данных не проверена (нужна маржинальная выгрузка, R6)."),
    }


def build_agencies_churn():
    """Отток/рост/миграции по агентствам — лист «3. Отток и миграции» (H1'25 vs H1'26 по
    транзакциям): сегменты Rising/Stable/Declining/Lost/New + список потерянных/новых с типом
    (тип — кураторская метка CHURN_TYPE_MAP, в самом листе её нет)."""
    src = find_agencies_source()
    if not src or "3. Отток и миграции" not in openpyxl.load_workbook(
            src, read_only=True).sheetnames:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    counts, detail, mode = {}, [], None
    for r in wb["3. Отток и миграции"].iter_rows(values_only=True):
        c0 = clean(r[0])
        if c0 == "Сегмент" and clean(r[1]) == "Агентств":
            mode = "counts"; continue
        if c0 == "Агентство" and clean(r[1]) == "H1-2025":
            mode = "detail"; continue
        if not c0:
            continue
        if mode == "counts" and c0 in ("Rising", "Stable", "Declining", "Lost", "New"):
            counts[c0] = int(_anum(r[1]) or 0)
        elif mode == "detail":
            detail.append({"name": c0, "h1_2025": _anum(r[1]), "h1_2026": _anum(r[2]),
                           "yoy": _anum(r[3]), "segment": clean(r[4])})

    lost = [{**d, "type": CHURN_TYPE_MAP.get(d["name"])} for d in detail if d["segment"] == "Lost"]
    return {
        "window": "H1 2025 (янв–июн) → H1 2026, по транзакциям",
        "counts": counts, "lost": lost,
        "newAgencies": [d for d in detail if d["segment"] == "New"],
    }


def build_agencies():
    """Свод по агентствам (срез) → agencies.json: на агентство + агрегаты направления."""
    src = find_agencies_source()
    if not src:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if s.lower().startswith("свод")), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    num = _anum

    # Месяц колонки «Оффлайн» берём из её заголовка: оффлайн по агентствам приходит из
    # транзакционного дампа, который отстаёт от общего отчёта. С июля 2026 месяц оффлайна
    # и месяц ops расходятся — подпись на странице обязана это показывать.
    offline_cut = None
    m_off = re.search(r"\(([^)]+)\)", str(rows[0][12] or "")) if len(rows[0]) > 12 else None
    if m_off:
        offline_cut = m_off.group(1).strip().lower()

    agencies, total = [], None
    for r in rows[1:]:
        name = clean(r[1])
        if not name:
            continue
        rec = {
            "name": name, "segment": clean(r[2]),
            "may": num(r[3]), "apr": num(r[4]), "momPct": num(r[5]),
            "l3m": num(r[6]), "l6m": num(r[7]), "may25": num(r[8]), "yoyPct": num(r[9]),
            "sharePct": num(r[10]),
            # Оффлайн (лист «Свод», колонки L/M): доля операций месяца среза, оформленных
            # вручную (не онлайн). ВАЖНО: с 2026-06 это снимок месяца среза, а не накопление
            # за 6 мес., как считалось раньше — базы разных месяцев напрямую не сравнивать.
            "offlineN": num(r[11]), "offlinePct": num(r[12]),
            "band": clean(r[13]),
        }
        if str(name).startswith("ИТОГО"):
            total = {"may": rec["may"], "apr": rec["apr"], "momPct": rec["momPct"], "count": len(agencies)}
            continue
        rec["id"] = clean(r[0])
        agencies.append(rec)

    # --- помесячные ряды по транзакциям (для графиков динамики и спарклайнов) ---
    # Только колонки вида «MM.YYYY» (служебные «Δ …» отбрасываем).
    def parse_monthly(sheet_name, resolve):
        rows_m = list(wb[sheet_name].iter_rows(values_only=True))
        hdr = [clean(x) for x in rows_m[0][1:]]
        keep = [i for i, h in enumerate(hdr) if h and re.match(r"^\d{2}\.\d{4}$", str(h))]
        months_m = [hdr[i] for i in keep]
        by_id_m, total_m = {}, None
        for r in rows_m[1:]:
            nm = clean(r[0])
            if not nm:
                continue
            def to_int(v):
                n = num(v)
                return int(round(n)) if n is not None else None
            series = [to_int(r[1 + i]) for i in keep]
            if str(nm).startswith("ИТОГО"):
                total_m = series
                continue
            aid = resolve(nm)
            if aid is not None:
                by_id_m[str(aid)] = series
        return {"months": months_m, "total": total_m, "byId": by_id_m}

    name2id_full = {a["name"]: a["id"] for a in agencies}
    mon_ops_sheet = next((s for s in wb.sheetnames if "операции l13m" in s.lower()), None)
    monthly = parse_monthly(mon_ops_sheet, name2id_full.get) if mon_ops_sheet else None

    # --- клиенты (NSM): история по месяцам разворачивается из «Общий отчёт …xlsx» (сырой дамп) ---
    cl = build_agencies_clients(agencies)
    if cl:
        for a in agencies:
            rec = cl["byId"].get(a["id"])
            if rec:
                a.update(rec)
        nsm_total, monthly_clients = cl["nsm"], cl["monthlyClients"]
    else:
        nsm_total, monthly_clients = None, None

    shares = sorted((a["sharePct"] or 0) for a in agencies)[::-1]
    top3 = round(sum(shares[:3]), 4)
    top5 = round(sum(shares[:5]), 4)

    # читаем мета-заметку из вкладки «Легенда» свода (если есть)
    note = None
    if "Легенда" in wb.sheetnames:
        for r in wb["Легенда"].iter_rows(values_only=True):
            if r and r[0] and "источник" in str(r[0]).lower():
                note = clean(r[1]); break

    # Дата среза выводится из имени листа «Свод (<месяц> <год>)» — раньше была зашита
    # строкой «30.06.2026» и молча устаревала при смене месяца.
    _MON = {"январ": (1, 31), "феврал": (2, 28), "март": (3, 31), "апрел": (4, 30),
            "ма": (5, 31), "июн": (6, 30), "июл": (7, 31), "август": (8, 31),
            "сентябр": (9, 30), "октябр": (10, 31), "ноябр": (11, 30), "декабр": (12, 31)}
    cut = "—"
    m = re.search(r"\(([^)]*?)\s+(\d{4})\)", sheet)
    if m:
        low = m.group(1).lower()
        for k, (mm, dd) in _MON.items():
            if low.startswith(k):
                cut = f"{dd:02d}.{mm:02d}.{m.group(2)}"
                break

    return {
        "source": src.name, "cut": cut, "offlineCut": offline_cut, "metric": "операции (ops)",
        "note": note, "total": total,
        "concentration": {"top3": top3, "top5": top5},
        "nsm": nsm_total,
        "monthly": monthly,
        "monthlyClients": monthly_clients,
        "agencies": agencies,
        "abcdx": build_agencies_abcdx(agencies),
        "price": build_agencies_price(agencies),
        "churn": build_agencies_churn(),
    }


def find_work_xlsx(needle):
    """Файл *.xlsx, чьё имя (NFC, lower) содержит подстроку needle.

    Возвращает НАИБОЛЬШЕЕ имя по сортировке — версии датируются в имени
    («Общий отчет 2026 07.xlsx» > «… 2026 06.xlsx»), поэтому это latest.
    Раньше возвращался первый попавшийся glob — порядок ФС недетерминирован,
    и после появления нового месяца сборка могла молча взять старый файл."""
    cands = [p for p in WORK_DIR.glob("*.xlsx")
             if needle in unicodedata.normalize("NFC", p.name).lower()]
    return sorted(cands, key=lambda p: unicodedata.normalize("NFC", p.name))[-1] if cands else None


# --- пересчёт истории Support-ratio по ТОЧНОЙ привязке (HDE + Свод) --------
# Старые теплокарта/тренд считались прежним методом привязки обращений и
# несопоставимы с baseline 13,6. Здесь числитель — сырой экспорт HDE (колонка
# «Агентство» + дата заявки), знаменатель — Свод («Операции помесячно L12M»).
# Малые агентства (Космос Тревел, Альбатрос) в временной срез не берём:
# месячный знаменатель <300 операций даёт шум, а не сигнал (см. основную таблицу).
SUP_HIST_MAP = {
    # имя в HDE «Агентство»: (имя в Своде «Операции помесячно», отображаемое имя)
    "IBC": ("IBC", "IBC"),
    "АТН": ("ATH", "АТН"),
    "Аэроглобус": ("Aeroglobus", "Аэроглобус"),
    "Корпорейт Тревел": ("Корпорэйт Трэвел", "Корпорейт Трэвел"),
    "Глобал Эир": ("Global Air", "Глобал Эир"),
    "ИНТЕРСИТИ СЕРВИС": ("KMP Group", "Интерсити (KMP)"),
    "Казтур": ("KazTour", "Казтур"),
    "ГСП": ("ГСП-Сервис", "ГСП"),
}
SUP_HIST_ORDER = ["IBC", "АТН", "Аэроглобус", "Корпорейт Тревел", "Глобал Эир",
                  "ИНТЕРСИТИ СЕРВИС", "Казтур", "ГСП"]


def recompute_support_history():
    """Теплокарта + тренд Support-ratio по точной привязке. None — если нет исходников."""
    import datetime
    f_hde = find_work_xlsx("hde_report")
    f_svod = find_work_xlsx("свод по агентствам")
    if not f_hde or not f_svod:
        return None

    # числитель: обращения помесячно по агентству (точная колонка «Агентство»)
    wb = openpyxl.load_workbook(f_hde, data_only=True, read_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.strip().lower().startswith("hde_report")), wb[wb.sheetnames[0]])
    it = ws.iter_rows(values_only=True)
    pos = {h: i for i, h in enumerate(next(it))}
    di, ai = pos.get("Дата создание заявки"), pos.get("Агентство")
    if di is None or ai is None:
        return None
    calls = {}
    for r in it:
        d, a = r[di], r[ai]
        if isinstance(d, datetime.datetime) and a:
            calls.setdefault(a, {})
            key = f"{d.year}-{d.month:02d}"
            calls[a][key] = calls[a].get(key, 0) + 1

    # знаменатель: операции помесячно
    wb2 = openpyxl.load_workbook(f_svod, data_only=True, read_only=True)
    ws2 = next((wb2[s] for s in wb2.sheetnames if "операции помесячно" in s.lower()), None)
    if not ws2:
        return None
    o = list(ws2.iter_rows(values_only=True))
    mcols = {}
    for i in range(1, len(o[0])):
        m = re.match(r"^(\d{2})\.(\d{4})$", str(o[0][i]).strip()) if o[0][i] else None
        if m:
            mcols[i] = f"{m.group(2)}-{m.group(1)}"
    months = [mcols[i] for i in sorted(mcols)]
    ops = {r[0]: {mcols[i]: r[i] for i in sorted(mcols)} for r in o[1:] if r[0]}

    def fnum(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None

    def slope(series):
        pts = [(i, v) for i, v in enumerate(series) if v is not None]
        n = len(pts)
        if n < 2:
            return None
        sx = sum(p[0] for p in pts); sy = sum(p[1] for p in pts)
        sxx = sum(p[0] ** 2 for p in pts); sxy = sum(p[0] * p[1] for p in pts)
        den = n * sxx - sx * sx
        return round((n * sxy - sx * sy) / den, 3) if den else None

    def mean(xs):
        xs = [x for x in xs if x is not None]
        return round(sum(xs) / len(xs), 2) if xs else None

    h_agencies, per = [], []
    for hde in SUP_HIST_ORDER:
        svod, disp = SUP_HIST_MAP[hde]
        if svod not in ops:
            continue
        vals = []
        for m in months:
            c = calls.get(hde, {}).get(m, 0)
            op = fnum(ops[svod].get(m))
            vals.append(round(c / op * 1000, 1) if op else None)
        h_agencies.append({"name": disp, "vals": vals})
        per.append({"name": disp, "slope": slope(vals),
                    "first6": mean(vals[:6]), "last6": mean(vals[-6:]),
                    "status": "спад ⤓" if (slope(vals) or 0) < -0.2 else "рост ⤒" if (slope(vals) or 0) > 0.2 else "плоско →"})

    total = []
    for m in months:
        sc = sum(calls.get(h, {}).get(m, 0) for h in SUP_HIST_ORDER)
        so = sum((fnum(ops[SUP_HIST_MAP[h][0]].get(m)) or 0) for h in SUP_HIST_ORDER if SUP_HIST_MAP[h][0] in ops)
        total.append(round(sc / so * 1000, 2) if so else None)

    tslope = slope(total)
    grow = sum(1 for p in per if (p["slope"] or 0) > 0)
    word = "растёт" if (tslope or 0) > 0 else "падает" if (tslope or 0) < 0 else "ровно"
    return {
        "heat": {"months": months, "agencies": h_agencies, "total": total},
        "trend": {
            "slopeText": f"{'+' if (tslope or 0) >= 0 else ''}{tslope} / мес ({word})",
            "slopeNum": tslope, "first6": mean(total[:6]), "last6": mean(total[-6:]),
            "conclusion": f"Пересчёт по точной привязке: ratio {word} ({'+' if (tslope or 0) >= 0 else ''}{tslope}/мес), {grow} из {len(per)} агентств не снижаются. Вывод «нагрузка не самоизлечивается» подтверждается на честном методе.",
            "perAgency": per,
            "method": "Точная привязка: числитель — HDE (колонка «Агентство»), знаменатель — Свод (операции помесячно). 8 агентств со стабильным знаменателем; малые (Космос Тревел, Альбатрос) исключены как шум.",
            "window": f"{months[0]} … {months[-1]}" if months else None,
        },
    }


def build_support():
    """Нагрузка на саппорт → support.json.

    Источники: «Обращения агентств — единый вывод …xlsx» (выводы, ratio, сегменты,
    категории, join ratio×NSM) + «Support-ratio baseline …xlsx» (помесячный ratio L13M,
    тренд). Ось «частота боли» (Impact), НЕ канал Reach. Числа — отсюда, не из HTML.
    """
    f1 = find_work_xlsx("единый вывод")
    f2 = find_work_xlsx("support-ratio baseline")
    if not f2:
        # Файл уехал в /Архив при уборке 2026-07-25 — сборка молча переставала
        # обновлять support.json, а манифест продолжал помечать его «generated».
        # Ищем в архиве и предупреждаем: данные живы, но источник вне корня.
        arch = sorted(WORK_DIR.glob("Архив/**/*.xlsx"))
        f2 = next((p for p in arch
                   if "support-ratio baseline" in unicodedata.normalize("NFC", p.name).lower()), None)
        if f2:
            print(f"  ⚠ support: источник тренда взят из архива — {f2.relative_to(WORK_DIR)}")
    if not f1 or not f2:
        print("  ⚠ support.json НЕ пересобран: не найден "
              + ("«Обращения агентств — единый вывод …xlsx»" if not f1 else "«Support-ratio baseline …xlsx»"))
        return None
    wb1 = openpyxl.load_workbook(f1, data_only=True, read_only=True)
    wb2 = openpyxl.load_workbook(f2, data_only=True, read_only=True)

    def sheet(wb, *needles):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(n.lower() in sl for n in needles):
                return wb[s]
        return None

    def num(v):
        if v is None or v == "":
            return None
        try:
            return round(float(v), 4)
        except (TypeError, ValueError):
            return None

    def rows_of(ws):
        return list(ws.iter_rows(values_only=True)) if ws else []

    def last_cell(r):
        """Последняя непустая ячейка строки (значения в сводных листах прижаты вправо)."""
        for v in reversed(r):
            c = clean(v)
            if c is not None:
                return c
        return None

    def digits(v):
        """Целое из строки с разделителями-пробелами (886, «9 271»)."""
        m = re.search(r"\d[\d\s ]*", str(v or ""))
        return int(re.sub(r"[\s ]", "", m.group())) if m else None

    # --- мета и метод (лист «0. Сводка и метод») ---
    data_period = rows_n = sup_share = sup_calls = sup_total = prev_ratio = method_note = None
    for r in rows_of(sheet(wb1, "сводка")):
        c0 = clean(r[0])
        if not c0:
            continue
        v = next((clean(x) for x in r[1:] if clean(x) is not None), None)
        low = c0.lower()
        if low.startswith("период данных"):
            data_period = re.sub(r"\s*\(.*\)\s*$", "", v).strip() if v else v
        elif low.startswith("строк"):
            rows_n = digits(v)
        elif low.startswith("было 19"):
            method_note = v
            pm = re.search(r"\d+[.,]\d+", str(v or ""))
            prev_ratio = float(pm.group().replace(",", ".")) if pm else None
        elif low.startswith("доля обращений к поставщику"):
            sm = re.search(r"(\d+[.,]\d+)\s*%", str(v or ""))
            sup_share = round(float(sm.group(1).replace(",", ".")) / 100, 4) if sm else None
            pair = re.search(r"\((\d[\d\s ]*)\s+из\s+(\d[\d\s ]*)\)", str(v or ""))
            if pair:
                sup_calls, sup_total = digits(pair.group(1)), digits(pair.group(2))

    # --- ratio по агентствам (лист «2. Support-ratio»): активные · BASELINE · ушедшие ---
    # Колонки: имя · сегмент · обращ.всего · обращ.L6M · опер.L6M · ratio · →поставщик · sup% · сигнал · боль.
    per_agency, per_agency_hist, baseline = [], [], None
    hist_mode = False
    for r in rows_of(sheet(wb1, "support-ratio"))[1:]:
        nm = clean(r[0])
        if not nm:
            continue
        if str(nm).startswith("Без операций"):
            hist_mode = True
            continue
        rec = {
            "name": nm, "segment": clean(r[1]),
            "callsTotal": num(r[2]), "callsL6M": num(r[3]), "ops": num(r[4]),
            "ratio": num(r[5]), "toSupplier": num(r[6]), "supShare": num(r[7]),
            "signal": clean(r[8]), "pain": clean(r[9]),
        }
        if str(nm).upper().startswith("BASELINE"):
            baseline = {"ratio": rec["ratio"], "calls": rec["callsL6M"], "ops": rec["ops"],
                        "toSupplier": rec["toSupplier"], "prevRatio": prev_ratio,
                        "methodNote": method_note}
            continue
        sig = (rec["signal"] or "")
        rec["smallBase"] = "мал. база" in sig
        rec["outflow"] = "отток" in sig
        (per_agency_hist if hist_mode else per_agency).append(rec)

    # --- помесячная динамика по агентствам (лист «5. Профиль агентств»): обращений в 2025 vs 2026 ---
    # Единственный временной срез новой (точной) атрибуции. 2026 — неполный год (до 24.06).
    prof = {}
    for r in rows_of(sheet(wb1, "профиль"))[1:]:
        nm = clean(r[0])
        if not nm:
            continue
        prof[nm] = {"y2025": num(r[2]), "y2026": num(r[3])}
    for rec in (*per_agency, *per_agency_hist):
        p = prof.get(rec["name"])
        if p:
            rec["y2025"], rec["y2026"] = p["y2025"], p["y2026"]

    # --- обращения к поставщику (лист «3. Обращения к поставщику»): блок A (агентства) + блок B (темы) ---
    sup_by_agency, sup_by_theme, mode = [], [], None
    for r in rows_of(sheet(wb1, "поставщик")):
        c0 = clean(r[0])
        if not c0:
            continue
        if c0.startswith("БЛОК A"):
            mode = "A"; continue
        if c0.startswith("БЛОК B"):
            mode = "B"; continue
        if c0 in ("Агентство", "Тема (L2)"):
            continue
        if mode == "A":
            sup_by_agency.append({"name": c0, "total": num(r[1]), "toSupplier": num(r[2]),
                                  "share": num(r[3]), "read": clean(r[4])})
        elif mode == "B":
            sup_by_theme.append({"theme": c0, "total": num(r[1]), "toSupplier": num(r[2]),
                                 "share": num(r[3]), "cls": clean(r[4])})

    # --- разбивка по темам (лист «4. Разбивка по темам») ---
    categories = []
    for r in rows_of(sheet(wb1, "разбивка"))[1:]:
        nm = clean(r[0])
        if not nm or num(r[1]) is None:
            continue
        categories.append({
            "theme": nm, "freq": num(r[1]), "shareBase": num(r[2]),
            "toSupplier": num(r[3]), "supShare": num(r[4]), "cls": clean(r[5]),
            "mapping": clean(r[6]),
        })
    categories.sort(key=lambda c: (c["freq"] or 0), reverse=True)

    # Чисто «Ракета-fixable» темы (0% поставщику) — кандидаты в селф-сервис №1.
    fixable = [{"theme": c["theme"], "freq": c["freq"]} for c in categories
               if c["cls"] and "fixable" in c["cls"].lower() and (c["toSupplier"] or 0) == 0]
    fixable.sort(key=lambda c: (c["freq"] or 0), reverse=True)

    supplier = {
        "share": sup_share, "calls": sup_calls, "total": sup_total or rows_n,
        "byAgency": sup_by_agency, "byTheme": sup_by_theme, "fixable": fixable[:6],
    }

    # --- сводка выводов (лист «1. Выводы»): # · вывод · что меняет ---
    # Статус для канона .ev-chip: вывод про пересборку baseline ссылается на открытый вопрос.
    conclusions = []
    for r in rows_of(sheet(wb1, "вывод"))[1:]:
        if num(r[0]) is None:
            continue
        n = int(num(r[0]))
        txt = clean(r[1])
        fact = clean(r[2])
        open_q = bool(txt and ("baseline" in txt.lower() or "пересчёт" in str(fact or "").lower()))
        conclusions.append({"n": n, "text": txt, "fact": fact,
                            "status": "Открыто" if open_q else "Подтверждено"})

    # --- помесячный ratio L13M (лист «Ratio×мес (L13M)») — теплокарта ---
    heat = None
    ws_h = sheet(wb2, "ratio")
    if ws_h:
        rows_h = rows_of(ws_h)
        hdr = rows_h[0]
        keep = [i for i, v in enumerate(hdr[1:]) if clean(v) and re.match(r"^\d{4}-\d{2}$", str(clean(v)))]
        months = [clean(hdr[1 + i]) for i in keep]
        h_agencies, h_total = [], None
        for r in rows_h[1:]:
            nm = clean(r[0])
            if not nm:
                continue
            if any(str(nm).startswith(p) for p in ("Метрика", "Цветовая", "Источник")):
                continue
            vals = [num(r[1 + i]) for i in keep]
            if str(nm).startswith("ИТОГО"):
                h_total = vals
                continue
            h_agencies.append({"name": nm, "vals": vals})
        heat = {"months": months, "agencies": h_agencies, "total": h_total}

    # --- тренд во времени (лист «Тренд Support-ratio — вывод») ---
    trend = {"slopeText": None, "slopeNum": None, "first6": None, "last6": None,
             "conclusion": None, "perAgency": []}
    for r in rows_of(sheet(wb2, "тренд")):
        c0 = clean(r[0])
        if not c0:
            continue
        low = c0.lower()
        if c0.startswith("ВЫВОД"):
            trend["conclusion"] = last_cell(r)
        elif "наклон" in low and "total" in low:
            txt = last_cell(r)
            trend["slopeText"] = txt
            m = re.search(r"[-+]?\d+[.,]?\d*", str(txt or ""))
            trend["slopeNum"] = float(m.group().replace(",", ".")) if m else None
        elif c0.startswith("Первые 6"):
            trend["first6"] = num(last_cell(r))
        elif c0.startswith("Последние 6"):
            trend["last6"] = num(last_cell(r))
        elif num(r[1]) is not None and not c0.startswith("Агентство"):
            trend["perAgency"].append({"name": c0, "slope": num(r[1]), "first6": num(r[2]),
                                       "last6": num(r[3]), "status": clean(r[4]), "read": clean(r[5])})

    # Пересчёт теплокарты/тренда по ТОЧНОЙ привязке (HDE + Свод) — заменяет
    # старый метод из baseline.xlsx, несопоставимый с 13,6. Если исходников нет —
    # остаётся старый расчёт из wb2 (graceful fallback).
    rec = recompute_support_history()
    if rec:
        heat = rec["heat"]
        trend = rec["trend"]

    # Доля 2026 года в данных (для пересчёта обращений в «в месяц»): по дате конца периода.
    months_2026 = None
    em = re.search(r"(\d{2})\.(\d{2})\.2026\s*$", str(data_period or ""))
    if em:
        months_2026 = round((int(em.group(2)) - 1) + int(em.group(1)) / 30.44, 2)

    return {
        "sourceCalls": f1.name, "sourceRatio": f2.name,
        "dataPeriod": data_period, "rows": rows_n, "months2026": months_2026,
        "metric": "Support-ratio — обращений в саппорт на 1000 операций (L6M: дек-25 … май-26)",
        "axisNote": "Ось «частота боли» (Impact), не канал Reach.",
        "baseline": baseline, "supplier": supplier,
        "trend": trend, "heat": heat,
        "perAgency": per_agency, "perAgencyHist": per_agency_hist,
        "categories": categories, "conclusions": conclusions,
        "windowShift": build_support_window_shift(per_agency, baseline),
    }


# Имя агентства в выгрузке обращений → имя в «Общем отчёте». Обращения приходят из HDE
# со своими написаниями, отчёт — из биллинга; общего ключа нет.
SUP_TO_REPORT = {
    "IBC": 'ООО "Интернейшнл Бизнес Центр" (IBC Corporate Travel)',
    "АТН": "ATH группа компаний", "Глобал Эир": "GLOBAL AIR", "ГСП": "ГСП-Сервис",
    "ИНТЕРСИТИ СЕРВИС": 'KMP Group (ООО "ИНТЕРСИТИ СЕРВИС")',
    "Корпорейт Тревел": 'ООО "КОРПОРЭЙТ ТРЭВЕЛ"', "Аэроглобус": "АЭРОГЛОБУС БИЗНЕС ТРЭВЕЛ",
    "Казтур": "KazTour Corporate", "Альбатрос": 'ООО "СБ "Альбатрос"',
    "Космос Тревел": "Космос Тревел", "Симпл Флайт": "Симпл Флайт",
    "Аэротон": "Аэротон Корпорейт Трэвел", "РейнаТур": "Рейна-Тур",
    "Атланта": "Атланта БТК", "Интоп": "Ин.Топ Консалтинг", "Иналекс": "Иналекс",
}


def build_support_window_shift(per_agency, baseline):
    """Проверка знаменателя Support-ratio и его эластичность к окну.

    Зачем. Ratio = обращения / операции. Числитель обновляется только новой выгрузкой
    обращений (HDE), знаменатель — каждым «Общим отчётом». Между обновлениями окна
    расходятся, и число на странице начинает двигаться по причинам, не связанным с
    саппортом. Здесь мы (1) сверяем зашитый знаменатель с фактом из отчёта за ТО ЖЕ
    окно — это ловит ошибку в данных; (2) считаем, каким стал бы ratio на актуальном
    L6M-окне при том же числителе — это НЕ новый ratio, а мера того, насколько число
    держится на знаменателе.

    Возвращает None, если «Общий отчёт …xlsx» не найден."""
    src = find_work_xlsx("общий отчет")
    if not src or not per_agency:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["All. Помесячно"]
    rows = list(ws.iter_rows(min_row=19, max_row=72, values_only=True))
    hdr = rows[0]
    idx = {str(h): i for i, h in enumerate(hdr) if h and re.match(r"^\d{6}$", str(h))}
    months = sorted(idx)
    if len(months) < 8:
        return None

    ops = {}
    for r in rows[1:]:
        nm = clean(r[1])
        if not nm:
            continue
        for m, i in idx.items():
            v = r[i] if i < len(r) else None
            if v:
                ops.setdefault(str(nm), {})[m] = ops.get(str(nm), {}).get(m, 0) + v

    # Окно числителя зашито в источнике обращений (дек-25 … май-26); окно «сейчас» —
    # последние 6 месяцев отчёта. Если источник обращений обновят, правится здесь.
    w_calls = ["202512", "202601", "202602", "202603", "202604", "202605"]
    w_now = months[-6:]
    if not set(w_calls) <= set(months):
        return None

    def total(nm, w):
        d = ops.get(nm, {})
        return sum(d.get(m, 0) for m in w)

    out, sum_stored, sum_fact, sum_now, sum_calls = [], 0, 0, 0, 0
    for a in per_agency:
        rep = SUP_TO_REPORT.get(a["name"])
        if not rep or rep not in ops:
            continue
        fact, now, calls = total(rep, w_calls), total(rep, w_now), a.get("callsL6M") or 0
        stored = a.get("ops") or 0
        sum_stored += stored; sum_fact += fact; sum_now += now; sum_calls += calls
        out.append({
            "name": a["name"], "opsStored": stored, "opsFact": fact, "opsNow": now,
            "ratio": a.get("ratio"),
            "ratioNow": round(calls / now * 1000, 1) if now else None,
            "opsDeltaPct": round(now / fact - 1, 4) if fact else None,
            "smallBaseNow": now < 1000,
        })
    for r in out:
        r["ratioDelta"] = (round(r["ratioNow"] - r["ratio"], 1)
                           if r["ratioNow"] is not None and r["ratio"] is not None else None)
    out.sort(key=lambda r: -abs(r["ratioDelta"] or 0))

    return {
        "callsWindow": "дек'25 – май'26", "opsWindow": f"{w_now[0]} – {w_now[-1]}",
        "opsWindowRu": f"{w_now[0][4:6]}.{w_now[0][:4]} – {w_now[-1][4:6]}.{w_now[-1][:4]}",
        "checkStored": sum_stored, "checkFact": sum_fact,
        "checkOk": abs(sum_stored - sum_fact) <= max(10, 0.001 * (sum_fact or 1)),
        "baselineRatio": (baseline or {}).get("ratio"),
        "baselineRatioNow": round(sum_calls / sum_now * 1000, 1) if sum_now else None,
        "rows": out,
    }


# --- нормализация доказательной базы (ТЗ 13) -----------------------------
# Единый словарь этапов. Любое значение в research/etapy/sootv обязано быть отсюда.
STAGE_VOCAB = ["1", "2A", "2B", "3", "4", "5", "клиент", "вне этапов"]

# Осознанные gap'ы H-кодов (преамбула Синтеза v2.1): гипотезы, под которые
# находок ещё нет намеренно, — линтер не считает их ошибкой, помечает «ожидаемо».
GAP_HCODES = {"H1.5", "H2.1-доп", "H3.1a", "H3.1b", "H3.5-канд", "H4.3-канд"}


def parse_stages(raw):
    """Строка этапа реестра → (массив этапов из словаря, флаг роли|None).
    "1/2B"→["1","2B"]; "3-4"→["3","4"]; "клиент/конс"→(["клиент"],"клиент/конс");
    "—"/пусто→["вне этапов"]."""
    if raw is None or str(raw).strip() in ("—", ""):
        return ["вне этапов"], None
    s = str(raw).strip()
    role = None
    if "конс" in s:                       # «клиент/конс» — клиентский этап, но роль = консультант
        role = "клиент/конс"
        s = re.sub(r"/?\s*конс\.?", "", s).strip(" /")
    parts = [p.strip() for p in re.split(r"[/\-]", s) if p.strip()]
    return (parts or ["вне этапов"]), role


def parse_hcodes(raw):
    """Строка H-кодов реестра → массив. "H1.3/H2.3"→["H1.3","H2.3"];
    "H2.4 (частично)"→["H2.4"]; "—"/«— (…)»→[]."""
    if raw is None or str(raw).strip().startswith("—"):
        return []
    codes, seen = [], set()
    for c in re.findall(r"H\d+\.\d+[a-zа-я]?", str(raw)):
        if c not in seen:
            seen.add(c); codes.append(c)
    return codes


def parse_block(raw):
    """Колонка «Источник» → код блока синтеза. "Синтез A"→"A"; "Блок1 …"→"1";
    "Синтез C / Блок4"→"C"; "Синтез / Боли-статус"→"—"."""
    if raw is None:
        return "—"
    s = str(raw).strip()
    m = re.match(r"Синтез\s+([A-ZА-Я])\b", s)
    if m:
        return m.group(1)
    m = re.match(r"Блок\s*(\d+)", s)
    if m:
        return m.group(1)
    return "—"


def research_anchor(fid):
    """Канон якоря находки: "A1"→"f-a1"; "4.2"→"f-4-2"; "7.1"→"f-7-1"."""
    s = re.sub(r"[./\\]+", "-", str(fid).strip().lower())
    return "f-" + s


def build_research():
    """Реестр исследований (90 строк) → research.json. Имена людей обезличиваем."""
    src = find_work_xlsx("реестр исследован")
    if not src:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if "реестр" in s.lower()), wb.sheetnames[0])
    rows = list(wb[sheet].iter_rows(values_only=True))
    hdr = [clean(c) or "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    def cell(r, name):
        i = idx.get(name)
        return clean(r[i]) if i is not None and i < len(r) else None

    findings = []
    for r in rows[1:]:
        fid = cell(r, "ID")
        finding = cell(r, "Боль / находка")
        if not fid and not finding:
            continue
        reach_raw = cell(r, "Подтв. агентства (Reach)") or ""
        # обезличиваем: убираем скобки с именами людей
        reach = re.sub(r"\s*\([^)]*\)", "", str(reach_raw)).strip(" ·,")
        agencies = [a.strip() for a in re.split(r"[,/+]| и ", reach) if a.strip() and a.strip() != "—"]
        stage_raw, src_raw = cell(r, "Этап"), cell(r, "Источник")
        stages, stage_role = parse_stages(stage_raw)
        findings.append({
            "id": fid, "theme": cell(r, "Тема"), "finding": finding,
            "role": cell(r, "Роль"), "reach": reach or None, "reachCount": len(agencies),
            "hypStatus": cell(r, "Статус гипотезы"), "qty": cell(r, "Кол. данные"),
            "stage": stage_raw, "mechanism": cell(r, "Механизм"),
            "hCode": cell(r, "H"), "jtbd": cell(r, "JTBD (кратко)"),
            "status": cell(r, "Статус"), "src": src_raw,
            # --- нормализованные поля (ТЗ 13) ---
            "anchor": research_anchor(fid) if fid else None,
            "stages": stages, "stageRole": stage_role,
            "hCodes": parse_hcodes(cell(r, "H")),
            "block": parse_block(src_raw),
        })

    # --- валидация уникальности якорей ---
    anchors, dup = {}, []
    for f in findings:
        a = f.get("anchor")
        if a is None:
            continue
        if a in anchors:
            dup.append(f"{a} (id {anchors[a]!r} и {f['id']!r})")
        anchors[a] = f["id"]
    if dup:
        print(f"⚠ Реестр: неуникальные якоря находок ({len(dup)}): " + "; ".join(dup))

    return {"source": src.name, "count": len(findings), "findings": findings}


def lint_links(research, sootv=None):
    """Линтер связей доказательной базы (ТЗ 13, п.3). Не валит сборку —
    печатает расхождения: стратегический документ должен видеть свои дыры."""
    print("\n── Линтер связей ──")
    etapy_path = DATA_DIR / "etapy.json"
    if not etapy_path.exists():
        print("  etapy.json не найден — пропуск проверки H-кодов и ссылок.")
        return
    etapy = json.loads(etapy_path.read_text(encoding="utf-8"))
    findings = (research or {}).get("findings", [])

    def base(code):
        """H-код к базовому виду для сверки: «H3.5-канд»→«H3.5», «H3.1a»→«H3.1»."""
        m = re.match(r"(H\d+\.\d+)", str(code))
        return m.group(1) if m else None

    # покрытие гипотез находками
    research_bases = {base(c) for f in findings for c in f.get("hCodes", []) if base(c)}
    missing, gaps = [], []
    for st in etapy.get("stages", []):
        for h in st.get("hypotheses", []):
            code = h.get("code")
            b = base(code)
            if b is None:                              # H-Супервизор и т.п. — без номера
                continue
            if b in research_bases:
                continue
            if h.get("candidate") or code in GAP_HCODES or "канд" in str(code) or "доп" in str(code):
                gaps.append(code)
            else:
                missing.append(code)
    if missing:
        print(f"  ⚠ H-коды без находки в реестре ({len(missing)}): {', '.join(missing)}")
    if gaps:
        print(f"  · осознанные gap (ожидаемо, {len(gaps)}): {', '.join(gaps)}")
    if not missing:
        print("  ✓ все «боевые» H-коды этапов имеют находку (или числятся в gap)")

    # ссылки research:<id> / planned:<id> из etapy.json (появятся в ТЗ 15/17)
    ids = {f["id"] for f in findings if f.get("id")}
    blob = json.dumps(etapy, ensure_ascii=False)
    bad_ref = sorted({m for m in re.findall(r"research:([^\"\s,\]]+)", blob) if m not in ids})
    if bad_ref:
        print(f"  ⚠ ссылки research:<id> на несуществующие находки: {', '.join(bad_ref)}")
    planned_path = DATA_DIR / "planned.json"
    planned_ids = set()
    if planned_path.exists():
        pj = json.loads(planned_path.read_text(encoding="utf-8"))
        planned_ids = {x.get("id") for x in pj.get("items", pj.get("planned", []))}
    bad_pl = sorted({m for m in re.findall(r"planned:([^\"\s,\]]+)", blob)
                     if m not in planned_ids})
    if bad_pl:
        where = "planned.json" if planned_path.exists() else "planned.json (нет файла)"
        print(f"  ⚠ ссылки planned:<id> вне {where}: {', '.join(bad_pl)}")

    # связи-доказательства гипотез (ТЗ 15): evidence[].ref и researchTodo[].planned
    def iter_hyps():
        for st in etapy.get("stages", []):
            yield from st.get("hypotheses", [])
        if etapy.get("crossLayer", {}).get("churn"):
            yield etapy["crossLayer"]["churn"]
    ev_bad_r, ev_bad_p, todo_bad_p = set(), set(), set()
    for h in iter_hyps():
        for e in h.get("evidence", []):
            if e.get("kind") == "research" and e.get("ref") not in ids:
                ev_bad_r.add(e.get("ref"))
            if e.get("kind") == "planned" and e.get("ref") not in planned_ids:
                ev_bad_p.add(e.get("ref"))
    for st in etapy.get("stages", []):
        for t in st.get("researchTodo", []):
            if t.get("planned") and t["planned"] not in planned_ids:
                todo_bad_p.add(t["planned"])
    if ev_bad_r:
        print(f"  ⚠ evidence kind=research на несуществующие находки: {', '.join(sorted(ev_bad_r))}")
    if ev_bad_p or todo_bad_p:
        print(f"  ⚠ evidence/researchTodo kind=planned вне planned.json: {', '.join(sorted(ev_bad_p | todo_bad_p))}")
    if not (ev_bad_r or ev_bad_p or todo_bad_p):
        n_ev = sum(len(h.get("evidence", [])) for h in iter_hyps())
        print(f"  ✓ связи-доказательства этапов: {n_ev} чипов, битых research:/planned: ссылок нет")

    # словарь этапов: research.stages из единого словаря
    bad_stage = sorted({s for f in findings for s in f.get("stages", [])
                        if s not in STAGE_VOCAB})
    if bad_stage:
        print(f"  ⚠ этапы вне словаря {STAGE_VOCAB}: {', '.join(bad_stage)}")

    # словарь этапов в матрице соответствия (sootv): stageNum из единого словаря
    if sootv:
        bad_sn = sorted({r.get("stageNum") for r in sootv.get("rows", [])
                         if r.get("stageNum") not in STAGE_VOCAB})
        if bad_sn:
            print(f"  ⚠ sootv: stageNum вне словаря: {', '.join(bad_sn)}")
        unmatched = sootv["count"] - sootv.get("painMatched", 0)
        print(f"  · sootv: кодов боли без находки в реестре {unmatched} из {sootv['count']} "
              f"(описательные коды вроде «M-блок/паритет» — ожидаемо)")

    # находки «вне этапов» — для дозаполнения Владом в xlsx
    outside = [f["id"] for f in findings if f.get("stages") == ["вне этапов"]]
    if outside:
        print(f"  · находок «вне этапов» ({len(outside)}), этап дозаполнит Влад в реестре:")
        print("    " + ", ".join(outside))


def build_sootv(research):
    """Матрица соответствия (этап→гипотеза→боль→итерация→метрика) → sootv.json (ТЗ 16).
    Цитаты с названиями агентств оставляем; имена людей в реестре уже обезличены.
    painCode мапится на якорь реестра, где находка существует."""
    src = find_work_xlsx("соответствие исследован")
    if not src or "Матрица" not in openpyxl.load_workbook(src, read_only=True).sheetnames:
        return None
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb["Матрица"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [clean(c) or "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    def cell(r, name):
        i = idx.get(name)
        return clean(r[i]) if i is not None and i < len(r) else None

    version = None
    if "Легенда" in wb.sheetnames:
        for r in wb["Легенда"].iter_rows(values_only=True):
            if r and r[0] and str(r[0]).strip().lower() == "версия":
                version = clean(r[1]); break

    def stage_main(raw):                       # «2A Сокращение…»→«2»; «1/2»→«1»
        m = re.match(r"\d+", str(raw or ""))
        return m.group(0) if m else "—"

    def split_plus(raw):                       # «H1.1+H1.2»→["H1.1","H1.2"]
        return [p.strip() for p in str(raw or "").split("+") if p.strip()] if raw else []

    def split_metrics(raw):                    # «↓ Parity gap; ↑ Adoption»→[…]
        return [p.strip() for p in re.split(r"[;·]", str(raw or "")) if p.strip()] if raw else []

    rids = {f["id"] for f in (research or {}).get("findings", []) if f.get("id")}
    rows_out, matched = [], 0
    for r in rows[1:]:
        if not any(r):
            continue
        pain_code = cell(r, "Код боли")
        anchor = research_anchor(pain_code) if pain_code in rids else None
        if anchor:
            matched += 1
        rows_out.append({
            "stage": cell(r, "Этап"), "stageNum": stage_main(cell(r, "Этап")),
            "block": cell(r, "Блок"), "hyps": split_plus(cell(r, "Гипотеза")),
            "hypStatus": cell(r, "Статус гипотезы"),
            "painCode": pain_code, "painAnchor": anchor,
            "pain": cell(r, "Суть боли"), "quote": cell(r, "Цитата / данные"),
            "mechanism": cell(r, "Механизм"), "iter": cell(r, "Итерация-кандидат"),
            "artifact": cell(r, "Артефакт (файл)"),
            "metricsActive": split_metrics(cell(r, "Активная метрика")),
            "metricsTarget": split_metrics(cell(r, "Целевая метрика")),
            "moscow": cell(r, "MoSCoW-ориентир"), "reach": cell(r, "Reach-сигнал"),
            "status": cell(r, "Статус"), "conflict": cell(r, "Конфликт / примечание"),
        })
    return {"source": src.name, "version": version, "count": len(rows_out),
            "painMatched": matched, "rows": rows_out}


def build_home(items, agencies, research):
    """Сводка для дашбордов главной (Проекции / Работа / Данные) → home.json.
    Цифры считаем здесь, а не хардкодим в HTML. Авторские слои (lestnica/concepts/
    metrics/exec) читаем из готовых data/*.json — это счётчики, не источник истины."""
    def load(name, default=None):
        p = DATA_DIR / name
        if not p.exists():
            return default
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return default

    # --- Работа: бэклог ---
    moscow_order = ["Must", "Should", "Could", "Won't"]
    moscow = {k: sum(1 for x in items if x.get("moscow") == k) for k in moscow_order}
    scored = sum(1 for x in items if isinstance(x.get("finalScore"), (int, float)))
    themes = {}
    for x in items:
        t = x.get("theme")
        if t:
            themes[t] = themes.get(t, 0) + 1
    themes_sorted = sorted(themes.items(), key=lambda kv: kv[1], reverse=True)
    top = max(
        (x for x in items if isinstance(x.get("finalScore"), (int, float))),
        key=lambda x: x["finalScore"], default=None,
    )
    work = {
        "total": len(items),
        "moscow": moscow,
        "scored": scored,
        "unscored": len(items) - scored,
        "topTitle": (top or {}).get("title"),
        "topFinal": (top or {}).get("finalScore"),
        "themes": [{"name": n, "count": c} for n, c in themes_sorted],
        "mustCount": moscow.get("Must", 0),
    }

    # --- Данные: агентства + исследования + метрики + рынок ---
    findings = (research or {}).get("findings", [])
    confirmed = sum(1 for f in findings if str(f.get("hypStatus") or "").startswith("подтв"))
    closed = sum(1 for f in findings if str(f.get("hypStatus") or "") == "закрыто")
    metrics = load("metrics.json", {})
    mlist = metrics.get("metrics", []) if isinstance(metrics, dict) else []
    data_block = {
        "agenciesActive": (agencies or {}).get("total", {}).get("count"),
        "opsMay": (agencies or {}).get("total", {}).get("may"),
        "momPct": (agencies or {}).get("total", {}).get("momPct"),
        "top3": (agencies or {}).get("concentration", {}).get("top3"),
        "top5": (agencies or {}).get("concentration", {}).get("top5"),
        "researchTotal": (research or {}).get("count"),
        "researchConfirmed": confirmed,
        "researchClosed": closed,
        "metricsActive": sum(1 for m in mlist if m.get("type") == "active"),
        "metricsTarget": sum(1 for m in mlist if m.get("type") == "target"),
        "metricsBaseline": sum(1 for m in mlist if m.get("baseline") or m.get("baselineNum")),
    }

    # --- Проекции: видение / лестница / концепции / дерево ---
    exec_data = load("exec.json", {})
    ladder = load("ladder.json", {})
    concepts = load("concepts.json", {})
    clist = concepts.get("concepts", []) if isinstance(concepts, dict) else []
    proj = {
        "goal": (exec_data.get("progress") or {}).get("goal"),
        "stages": len(ladder.get("stages", [])) if isinstance(ladder, dict) else None,
        "stagesFocus": "1–2",
        "subgoals": len(ladder.get("subgoalMap", [])) if isinstance(ladder, dict) else None,
        "concepts": len(clist),
        "conceptsFocus": sum(1 for c in clist if c.get("q2") == "focus"),
        "conceptsPartial": sum(1 for c in clist if c.get("q2") == "partial"),
        "conceptsOut": sum(1 for c in clist if c.get("q2") == "out"),
        "treeRoles": 3,  # JTBD: консультант / супервизор / руководитель (редакторская онтология)
        "betsNames": [b.get("name") for b in exec_data.get("bets", [])],
    }

    return {
        "_meta": "Сводка дашбордов главной. Считается build.py из backlog/agencies/research + авторских JSON.",
        "projections": proj,
        "work": work,
        "data": data_block,
    }


def metric_anchor(code):
    """Зеркало slug метрики из app.js: «TTFO»→«m-ttfo», «T-op»→«m-t-op»."""
    s = re.sub(r"[^0-9a-zа-яё]+", "-", str(code).lower(), flags=re.I).strip("-")
    return "m-" + s


def _snippet(text, limit=140):
    t = " ".join(str(text or "").split())
    return t[: limit - 1] + "…" if len(t) > limit else t


def build_search(items, research, tree):
    """Лёгкий индекс для глобального поиска (ТЗ 08): находки · метрики · легенды · темы.
    Агентства не индексируем (приватные имена/выручка). Формат: {title, page, href, snippet, kind}."""
    idx = []

    # Темы направления (из дерева) → карта уровней
    for t in (tree or []):
        theme = t.get("theme")
        if not theme:
            continue
        idx.append({
            "title": theme, "page": "Темы", "kind": "тема",
            "href": f"levels.html?theme={quote(theme)}",
            "snippet": _snippet(f"{t.get('count', 0)} итераций, Must {t.get('mustCount', 0)}"),
        })

    # Находки исследований → research.html#f-<id>
    for f in ((research or {}).get("findings") or []):
        fid = f.get("id")
        if not fid:
            continue
        idx.append({
            "title": f"{fid} · {_snippet(f.get('finding'), 60)}", "page": "Исследования", "kind": "находка",
            "href": f"research.html#{research_anchor(fid)}",
            "snippet": _snippet(f.get("finding")),
        })

    # Метрики → metrics.html#m-<code>
    metrics = _load_data_json("metrics.json", {})
    for m in (metrics.get("metrics") or []):
        code = m.get("code")
        if not code:
            continue
        idx.append({
            "title": f"{code} — {m.get('name', '')}".strip(" —"), "page": "Метрики", "kind": "метрика",
            "href": f"metrics.html#{metric_anchor(code)}",
            "snippet": _snippet(m.get("measures")),
        })

    # Легенда (определения) → legend.html#<section.id>
    legend = _load_data_json("legend.json", {})
    for sec in (legend.get("sections") or []):
        sid = sec.get("id")
        for it in (sec.get("items") or []):
            term = it.get("term")
            if not term:
                continue
            anchor = it.get("id") or sid
            idx.append({
                "title": term, "page": "Легенды", "kind": "термин",
                "href": f"legend.html#{anchor}" if anchor else "legend.html",
                "snippet": _snippet(it.get("def")),
            })

    return idx


# Данные сайта делятся на два класса, и «дата данных» у них берётся по-разному.
# GENERATED пересобираются из .xlsx этим скриптом — их дата = дата сборки.
# Остальные json авторские: правятся руками и `make build` их не трогает,
# поэтому они тихо стареют — их дата = mtime файла.
# etapy.json, planned.json, levels.json сюда НЕ входят: main() их только
# читает (линтер связей), но не пересобирает — де-факто они авторские.
GENERATED_JSON = {
    "backlog.json", "initiatives.json", "tree.json", "agencies.json",
    "support.json", "research.json", "sootv.json", "legend.json",
    "home.json", "search.json",
}

# Чем питается каждый генерируемый json — чтобы в манифесте была видна не
# только дата, но и происхождение числа.
JSON_SOURCE_HINT = {
    "backlog.json": "единый бэклог",
    "tree.json": "единый бэклог",
    "legend.json": "единый бэклог, вкладка «Легенда»",
    "search.json": "производный индекс (бэклог + реестр + дерево)",
    "home.json": "производная сводка (бэклог + агентства + реестр)",
    "initiatives.json": "сокращённый бэклог (инициативы)",
    "agencies.json": "«Свод по агентствам …xlsx» + «Общий отчёт …xlsx»",
    "support.json": "«Обращения агентств — единый вывод …xlsx»",
    "research.json": "«Реестр исследований …xlsx»",
    "sootv.json": "«Соответствие исследований и логики ценности.xlsx»",
}

STALE_AFTER_DAYS = 30  # порог «данные устарели» для плашки в футере


def _manifest_source(name, sources):
    """Человекочитаемый источник json: где можно — с именем файла-донора."""
    hint = JSON_SOURCE_HINT.get(name)
    if hint is None:
        return "пересобирается `make build`"
    return sources.get(hint, hint)


def build_manifest(sources):
    """data/manifest.json — дата и происхождение каждого файла данных.

    Заменяет ручную константу SITE_UPDATED в app.js: одна дата на весь сайт,
    выставленная руками, устаревала незаметно и подписывала июньскими
    данными июльские страницы (и наоборот).
    """
    from datetime import datetime, timezone

    def iso(ts):
        return datetime.fromtimestamp(ts, timezone.utc).astimezone().isoformat(timespec="seconds")

    files = {}
    for p in sorted(DATA_DIR.glob("*.json")):
        if p.name == "manifest.json":
            continue
        generated = p.name in GENERATED_JSON
        files[p.name] = {
            "kind": "generated" if generated else "authored",
            "date": iso(p.stat().st_mtime),
            "source": _manifest_source(p.name, sources) if generated
                      else "авторский файл, правится вручную",
        }
    return {
        "builtAt": build_meta()["builtAt"],
        "staleAfterDays": STALE_AFTER_DAYS,
        "sources": sources,
        "files": files,
    }


def _load_data_json(name, default):
    p = DATA_DIR / name
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return default


def main():
    print(f"Скоринг: {SCORING_PATH.name} (сверен с каноном {SCORING.get('synced')})")
    items, errors = build_backlog()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out = DATA_DIR / "backlog.json"
    payload = {
        "source": find_source().name,
        "count": len(items),
        "meta": build_meta(),
        "columns": list(COLUMNS.values()),
        "items": items,
    }
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Записано {len(items)} итераций → {out.relative_to(SITE_DIR)}")

    # initiatives.json — сокращённый бэклог (укрупнённые инициативы), отдельный источник
    initiatives, init_errors = build_initiatives()
    if initiatives:
        init_out = DATA_DIR / "initiatives.json"
        init_out.write_text(json.dumps(initiatives, ensure_ascii=False, indent=2), encoding="utf-8")
        imust = sum(1 for x in initiatives["items"] if x.get("moscow") == "Must")
        print(f"Инициативы: {initiatives['count']} (Must {imust}) → {init_out.relative_to(SITE_DIR)}")
    errors = errors + [f"[инициативы] {m}" for m in init_errors]

    # tree.json — дерево из того же источника
    tree = build_tree(items)
    tree_out = DATA_DIR / "tree.json"
    tree_out.write_text(json.dumps({
        "source": find_source().name,
        "meta": build_meta(),
        "count": len(items),
        "themes": len(tree),
        "tree": tree,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Дерево: {len(tree)} тем × Уровень 2 × итерации → {tree_out.relative_to(SITE_DIR)}")

    # legend.json — справочник из вкладки «Легенда»
    legend = build_legend(find_source())
    if legend:
        legend_out = DATA_DIR / "legend.json"
        legend_out.write_text(json.dumps(legend, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Легенда: {len(legend['sections'])} секций · {legend['count']} определений → {legend_out.relative_to(SITE_DIR)}")

    # agencies.json — свод по агентствам
    agencies = build_agencies()
    if agencies:
        ag_out = DATA_DIR / "agencies.json"
        ag_out.write_text(json.dumps(agencies, ensure_ascii=False, indent=2), encoding="utf-8")
        c = agencies["concentration"]
        print(f"Агентства: {len(agencies['agencies'])} активных · ТОП-3 {c['top3']:.0%} / ТОП-5 {c['top5']:.0%} → {ag_out.relative_to(SITE_DIR)}")

    # support.json — нагрузка на саппорт (Support-ratio)
    support = build_support()
    if support:
        sup_out = DATA_DIR / "support.json"
        sup_out.write_text(json.dumps(support, ensure_ascii=False, indent=2), encoding="utf-8")
        b = support.get("baseline") or {}
        sup = support.get("supplier") or {}
        print(f"Саппорт: {len(support['perAgency'])} активных + {len(support['perAgencyHist'])} ушедших · "
              f"baseline {b.get('ratio')} (было {b.get('prevRatio')}) · поставщик {sup.get('calls')} ({sup.get('share')}) · "
              f"{len(support['conclusions'])} выводов · {len(support['categories'])} тем "
              f"→ {sup_out.relative_to(SITE_DIR)}")

    # research.json — реестр исследований
    research = build_research()
    if research:
        r_out = DATA_DIR / "research.json"
        r_out.write_text(json.dumps(research, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Исследования: {research['count']} находок → {r_out.relative_to(SITE_DIR)}")

    # sootv.json — матрица соответствия (доказательная база, ТЗ 16)
    sootv = build_sootv(research)
    if sootv:
        s_out = DATA_DIR / "sootv.json"
        s_out.write_text(json.dumps(sootv, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Соответствие: {sootv['count']} строк матрицы · кодов боли с находкой "
              f"{sootv['painMatched']} → {s_out.relative_to(SITE_DIR)}")

    if research:
        lint_links(research, sootv)

    # home.json — сводка дашбордов главной
    home = build_home(items, agencies, research)
    home_out = DATA_DIR / "home.json"
    home_out.write_text(json.dumps(home, ensure_ascii=False, indent=2), encoding="utf-8")
    w, dd = home["work"], home["data"]
    print(f"Главная: бэклог {w['total']} (Must {w['mustCount']}) · агентства {dd['agenciesActive']} · находки {dd['researchTotal']} → {home_out.relative_to(SITE_DIR)}")

    # search.json — лёгкий индекс глобального поиска (ТЗ 08)
    search = build_search(items, research, tree)
    search_out = DATA_DIR / "search.json"
    search_out.write_text(json.dumps({"count": len(search), "index": search}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Поиск: индекс {len(search)} записей → {search_out.relative_to(SITE_DIR)}")

    must = sum(1 for x in items if x.get("moscow") == "Must")
    print(f"  Must: {must}  ·  с Final Score: {sum(1 for x in items if x['finalScore'] is not None)}")

    # Единый термин по всем собранным json (устойчиво к пересборке).
    normalize_terms()

    # manifest.json — последним: normalize_terms() трогает mtime файлов.
    sources = {"единый бэклог": find_source().name}
    init_src = find_initiatives_source()
    if init_src:
        sources["сокращённый бэклог (инициативы)"] = init_src.name
    manifest = build_manifest(sources)
    (DATA_DIR / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    authored = [n for n, f in manifest["files"].items() if f["kind"] == "authored"]
    print(f"Манифест: {len(manifest['files'])} файлов данных "
          f"({len(authored)} авторских) → data/manifest.json")

    if errors:
        print(f"\n⚠ Предупреждения валидации ({len(errors)}):")
        # Сводка по классам — чтобы из лога было видно, ЧТО чинить,
        # а не только сколько строк не нравится валидатору.
        classes = (
            ("обрезана при записи", "обрезанная ячейка — текст потерян в источнике"),
            ("связь §4 не сходится", "этап ↔ подцель не сходятся (§4)"),
            ("без кода H1.x", "гипотеза без кода (§5.3)"),
            ("составное значение", "составной этап (§5.1 требует одного)"),
            ("Концентрация", "Концентрация вне словаря §5.4"),
            ("Подцель", "Подцель вне словаря §4"),
            ("Механизм", "Механизм вне словаря §5.2"),
            ("Блок", "Блок вне словаря"),
            ("Этап", "Этап вне словаря §5.1"),
            ("пустая подцель", "пустая подцель"),
            ("пустая гипотеза", "пустая гипотеза"),
            ("нет в scoring.yml", "неизвестный H-код"),
        )
        tally = {}
        for e in errors:
            for pat, name in classes:
                if pat in e:
                    tally[name] = tally.get(name, 0) + 1
                    break
            else:
                tally["прочее"] = tally.get("прочее", 0) + 1
        for name, n in sorted(tally.items(), key=lambda kv: -kv[1]):
            print(f"  {n:>4}  {name}")
        report = DATA_DIR.parent / "build" / "validation-errors.txt"
        report.write_text(
            "# Ошибки валидации бэклога и инициатив.\n"
            "# Пересобирается каждым build.py. Порядок — как в источнике.\n\n"
            + "\n".join(errors) + "\n", encoding="utf-8")
        print(f"  полный список → site/build/{report.name}")
    else:
        print("Валидация: ошибок нет.")

    if GAP_NOTES:
        by_gap = {}
        for msg in GAP_NOTES:
            key = msg.split(": ", 1)[1]
            by_gap[key] = by_gap.get(key, 0) + 1
        print(f"\n· Осознанные gap'ы гипотез ({len(GAP_NOTES)} строк) — "
              f"код не назначен намеренно, решение зафиксировано в scoring.yml:")
        for key, n in sorted(by_gap.items(), key=lambda kv: -kv[1]):
            print(f"  · {n:>3}  {key}")

    # Написания вне канона — отдельным блоком: это не поломанные данные,
    # а очередь на нормализацию. Полный список пишется в файл, чтобы его
    # можно было отрабатывать построчно, а не выуживать из лога.
    if STYLE_NOTES:
        print(f"\n· Написания вне канона ({len(STYLE_NOTES)}) — данные читаются, "
              f"но словарь §4/§5.1/§5.2/§5.4 записан не канонически:")
        for msg in STYLE_NOTES[:10]:
            print("  ·", msg)
        if len(STYLE_NOTES) > 10:
            print(f"  … ещё {len(STYLE_NOTES) - 10}")
        report = DATA_DIR.parent / "build" / "normalize-queue.txt"
        report.write_text(
            "# Очередь нормализации. Пересобирается каждым build.py.\n"
            "# Опознанные значения, записанные не по словарю канона.\n\n"
            + "\n".join(STYLE_NOTES) + "\n", encoding="utf-8")
        print(f"  полный список → site/build/{report.name}")


if __name__ == "__main__":
    main()
